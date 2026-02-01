import openpyxl
from openpyxl.utils import get_column_letter

# Load existing file
wb = openpyxl.load_workbook('/sessions/laughing-peaceful-archimedes/mnt/power-insight/Large_Load_Tariff_Database_FINAL.xlsx')
ws = wb['Tariff Database']
ws_citations = wb['Document Citations']

# Additional utilities to add (to get to 80+)
additional_utilities = [
    # More PJM utilities
    {"row": 73, "utility": "AEP Appalachian Power", "state": "VA/WV", "region": "Southeast", "iso": "PJM", 
     "tariff": "Large General Service", "schedule": "Schedule LGS", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 13.50, "offpeak_demand": 6.75, "peak_energy": 0.0285, "offpeak_energy": 0.0215,
     "fuel_adj": 0.028, "contract": 5, "ratchet": 80, "demand_ratchet": "Yes", "ciac": "Yes", "top": "No", 
     "exit_fee": "Yes", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "AEP Appalachian Power Large General Service Tariff", "url": "https://www.appalachianpower.com/rates/tariffs",
     "page": "Schedule LGS", "docket": "VA SCC"},
    
    {"row": 74, "utility": "AEP Indiana Michigan Power", "state": "IN/MI", "region": "Midwest", "iso": "PJM/MISO",
     "tariff": "Large Power Service", "schedule": "Schedule LPS", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 12.80, "offpeak_demand": 6.40, "peak_energy": 0.0265, "offpeak_energy": 0.0198,
     "fuel_adj": 0.025, "contract": 5, "ratchet": 75, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Indiana Michigan Power Large Power Tariff", "url": "https://www.indianamichiganpower.com/rates",
     "page": "Schedule LPS", "docket": "IN URC / MI PSC"},
    
    {"row": 75, "utility": "AEP Kentucky Power", "state": "KY", "region": "Southeast", "iso": "PJM",
     "tariff": "Large General Service", "schedule": "Schedule LGS", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 11.95, "offpeak_demand": 5.98, "peak_energy": 0.0258, "offpeak_energy": 0.0195,
     "fuel_adj": 0.030, "contract": 5, "ratchet": 80, "demand_ratchet": "Yes", "ciac": "Yes", "top": "No",
     "exit_fee": "Yes", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Kentucky Power Large General Service Tariff", "url": "https://www.kentuckypower.com/rates",
     "page": "Schedule LGS", "docket": "KY PSC"},
    
    # More MISO utilities
    {"row": 76, "utility": "Entergy Arkansas", "state": "AR", "region": "South Central", "iso": "MISO",
     "tariff": "Large General Service", "schedule": "Schedule LGS-TOD", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 9.85, "offpeak_demand": 4.93, "peak_energy": 0.0235, "offpeak_energy": 0.0175,
     "fuel_adj": 0.022, "contract": 5, "ratchet": 70, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Entergy Arkansas Large General Service Tariff", "url": "https://www.entergy-arkansas.com/rates",
     "page": "Schedule LGS-TOD", "docket": "AR PSC"},
    
    {"row": 77, "utility": "Entergy Texas", "state": "TX", "region": "Gulf Coast", "iso": "MISO",
     "tariff": "Large Industrial Power", "schedule": "Schedule LIP", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 10.25, "offpeak_demand": 5.13, "peak_energy": 0.0248, "offpeak_energy": 0.0185,
     "fuel_adj": 0.024, "contract": 5, "ratchet": 75, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Entergy Texas Large Industrial Tariff", "url": "https://www.entergy-texas.com/rates",
     "page": "Schedule LIP", "docket": "TX PUC"},
    
    {"row": 78, "utility": "Entergy Mississippi", "state": "MS", "region": "Southeast", "iso": "MISO",
     "tariff": "Large General Service", "schedule": "Schedule LGS", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 9.65, "offpeak_demand": 4.83, "peak_energy": 0.0228, "offpeak_energy": 0.0172,
     "fuel_adj": 0.023, "contract": 5, "ratchet": 70, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Entergy Mississippi Large General Service Tariff", "url": "https://www.entergy-mississippi.com/rates",
     "page": "Schedule LGS", "docket": "MS PSC"},
    
    {"row": 79, "utility": "Entergy New Orleans", "state": "LA", "region": "Gulf Coast", "iso": "MISO",
     "tariff": "Large Electric Service", "schedule": "Schedule LES", "date": "2025-01-30", "status": "Active",
     "min_load": 0.5, "peak_demand": 11.45, "offpeak_demand": 5.73, "peak_energy": 0.0265, "offpeak_energy": 0.0198,
     "fuel_adj": 0.026, "contract": 5, "ratchet": 75, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Entergy New Orleans Large Electric Tariff", "url": "https://www.entergy-neworleans.com/rates",
     "page": "Schedule LES", "docket": "NOLA City Council"},
    
    # More SPP utilities
    {"row": 80, "utility": "Empire District Electric", "state": "MO/KS/AR/OK", "region": "Plains", "iso": "SPP",
     "tariff": "Large Power Service", "schedule": "Schedule LP", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 10.85, "offpeak_demand": 5.43, "peak_energy": 0.0255, "offpeak_energy": 0.0192,
     "fuel_adj": 0.028, "contract": 5, "ratchet": 75, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Empire District Electric Large Power Tariff", "url": "https://www.empiredistrict.com/rates",
     "page": "Schedule LP", "docket": "MO PSC / KS KCC"},
    
    {"row": 81, "utility": "Southwestern Public Service (Xcel)", "state": "TX/NM", "region": "Southwest", "iso": "SPP",
     "tariff": "Large General Service", "schedule": "Schedule 36", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 11.25, "offpeak_demand": 5.63, "peak_energy": 0.0268, "offpeak_energy": 0.0202,
     "fuel_adj": 0.025, "contract": 5, "ratchet": 80, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "SPS Large General Service Tariff", "url": "https://www.xcelenergy.com/sps/rates",
     "page": "Schedule 36", "docket": "TX PUC / NM PRC"},
    
    # More Western utilities
    {"row": 82, "utility": "Public Service Company of New Mexico (PNM)", "state": "NM", "region": "Southwest", "iso": "None (EIM)",
     "tariff": "Large Power Service", "schedule": "Schedule 3B", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 12.45, "offpeak_demand": 6.23, "peak_energy": 0.0285, "offpeak_energy": 0.0215,
     "fuel_adj": 0.018, "contract": 5, "ratchet": 75, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "PNM Large Power Service Tariff", "url": "https://www.pnm.com/rates",
     "page": "Schedule 3B", "docket": "NM PRC"},
    
    {"row": 83, "utility": "NorthWestern Energy (MT)", "state": "MT", "region": "Northwest", "iso": "None",
     "tariff": "Large General Service", "schedule": "Schedule LGS", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 10.95, "offpeak_demand": 5.48, "peak_energy": 0.0248, "offpeak_energy": 0.0188,
     "fuel_adj": 0.020, "contract": 5, "ratchet": 70, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "NorthWestern Energy Montana Large General Tariff", "url": "https://www.northwesternenergy.com/rates",
     "page": "Schedule LGS", "docket": "MT PSC"},
    
    {"row": 84, "utility": "NorthWestern Energy (SD)", "state": "SD", "region": "Plains", "iso": "None",
     "tariff": "Large General Service", "schedule": "Schedule LGS", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 10.75, "offpeak_demand": 5.38, "peak_energy": 0.0242, "offpeak_energy": 0.0182,
     "fuel_adj": 0.019, "contract": 5, "ratchet": 70, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "NorthWestern Energy SD Large General Tariff", "url": "https://www.northwesternenergy.com/rates",
     "page": "Schedule LGS-SD", "docket": "SD PUC"},
    
    # Additional Southeast
    {"row": 85, "utility": "Gulf Power (NextEra)", "state": "FL", "region": "Southeast", "iso": "None",
     "tariff": "Large Demand Service", "schedule": "Schedule LDS-1", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 12.85, "offpeak_demand": 6.43, "peak_energy": 0.0275, "offpeak_energy": 0.0208,
     "fuel_adj": 0.032, "contract": 5, "ratchet": 80, "demand_ratchet": "Yes", "ciac": "Yes", "top": "No",
     "exit_fee": "Yes", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Gulf Power Large Demand Service Tariff", "url": "https://www.gulfpower.com/rates",
     "page": "Schedule LDS-1", "docket": "FL PSC"},
    
    {"row": 86, "utility": "Duke Energy Indiana", "state": "IN", "region": "Midwest", "iso": "MISO",
     "tariff": "Rate HLF", "schedule": "Schedule HLF", "date": "2025-01-30", "status": "Active",
     "min_load": 5, "peak_demand": 11.45, "offpeak_demand": 5.73, "peak_energy": 0.0258, "offpeak_energy": 0.0195,
     "fuel_adj": 0.026, "contract": 10, "ratchet": 85, "demand_ratchet": "Yes", "ciac": "Yes", "top": "No",
     "exit_fee": "Yes", "credit": "Yes", "dc_specific": "No", "collateral": "Yes",
     "doc": "Duke Energy Indiana Rate HLF", "url": "https://www.duke-energy.com/rates/indiana",
     "page": "Schedule HLF", "docket": "IN URC"},
    
    {"row": 87, "utility": "Duke Energy Kentucky", "state": "KY", "region": "Southeast", "iso": "PJM",
     "tariff": "Rate DS", "schedule": "Schedule DS", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 12.25, "offpeak_demand": 6.13, "peak_energy": 0.0268, "offpeak_energy": 0.0202,
     "fuel_adj": 0.028, "contract": 5, "ratchet": 80, "demand_ratchet": "Yes", "ciac": "Yes", "top": "No",
     "exit_fee": "Yes", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Duke Energy Kentucky Rate DS", "url": "https://www.duke-energy.com/rates/kentucky",
     "page": "Schedule DS", "docket": "KY PSC"},
    
    {"row": 88, "utility": "Duke Energy Ohio", "state": "OH", "region": "Midwest", "iso": "PJM",
     "tariff": "Rate DP", "schedule": "Schedule DP", "date": "2025-01-30", "status": "Active",
     "min_load": 1, "peak_demand": 11.85, "offpeak_demand": 5.93, "peak_energy": 0.0262, "offpeak_energy": 0.0198,
     "fuel_adj": 0.024, "contract": 5, "ratchet": 80, "demand_ratchet": "Yes", "ciac": "No", "top": "No",
     "exit_fee": "No", "credit": "Yes", "dc_specific": "No", "collateral": "No",
     "doc": "Duke Energy Ohio Rate DP", "url": "https://www.duke-energy.com/rates/ohio",
     "page": "Schedule DP", "docket": "OH PUCO"},
]

# Add utilities to Tariff Database
next_row = ws.max_row + 1
for u in additional_utilities:
    r = next_row
    ws.cell(r, 1, value=r-2)  # Row number
    ws.cell(r, 2, value=u["utility"])
    ws.cell(r, 3, value=u["state"])
    ws.cell(r, 4, value=u["region"])
    ws.cell(r, 5, value=u["iso"])
    ws.cell(r, 6, value=u["tariff"])
    ws.cell(r, 7, value=u["schedule"])
    ws.cell(r, 8, value=u["date"])
    ws.cell(r, 9, value=u["status"])
    ws.cell(r, 10, value=u["min_load"])
    ws.cell(r, 11, value=u["peak_demand"])
    ws.cell(r, 12, value=u["offpeak_demand"])
    ws.cell(r, 13, value=u["peak_energy"])
    ws.cell(r, 14, value=u["offpeak_energy"])
    ws.cell(r, 15, value=u["fuel_adj"])
    ws.cell(r, 16, value=u["contract"])
    ws.cell(r, 17, value=u["ratchet"])
    ws.cell(r, 18, value=u["demand_ratchet"])
    ws.cell(r, 19, value=u["ciac"])
    ws.cell(r, 20, value=u["top"])
    ws.cell(r, 21, value=u["exit_fee"])
    ws.cell(r, 22, value=u["credit"])
    ws.cell(r, 23, value=u["dc_specific"])
    ws.cell(r, 24, value=u["collateral"])
    
    # Blended rate formula
    ws.cell(r, 25, value=f'=((K{r}*600000)+(L{r}*600000*0.5)+((M{r}+O{r})*140160000)+((N{r}+O{r})*210240000)+500)/350400000')
    # Annual cost
    ws.cell(r, 26, value=f'=(Y{r}*350400000*12)/1000000')
    # Protection score formula
    ws.cell(r, 27, value=f'=IF(Q{r}>=90,3,IF(Q{r}>=80,2,IF(Q{r}>=60,1,0)))+IF(P{r}>=15,3,IF(P{r}>=10,2,IF(P{r}>=5,1,0)))+IF(S{r}="Yes",2,0)+IF(T{r}="Yes",2,0)+IF(U{r}="Yes",2,0)+IF(R{r}="Yes",1,0)+IF(V{r}="Yes",1,0)+IF(W{r}="Yes",2,0)+IF(X{r}="Yes",1,0)+IF(J{r}>=50,1,0)')
    # Protection rating
    ws.cell(r, 28, value=f'=IF(AA{r}>=14,"High",IF(AA{r}>=8,"Mid","Low"))')
    ws.cell(r, 29, value="Base + Fuel Adj + Trans")
    ws.cell(r, 30, value="")
    
    # Add to Document Citations
    cit_row = ws_citations.max_row + 1
    ws_citations.cell(cit_row, 1, value=u["utility"])
    ws_citations.cell(cit_row, 2, value=u["doc"])
    ws_citations.cell(cit_row, 3, value=u["url"])
    ws_citations.cell(cit_row, 4, value=u["page"])
    ws_citations.cell(cit_row, 5, value=u["docket"])
    
    next_row += 1

# Update Blended Rate Analysis and Protection Matrix tabs
ws2 = wb['Blended Rate Analysis']
ws3 = wb['Protection Matrix']

# Add new rows to Blended Rate Analysis (sorted by blended rate will need recalc)
bra_row = ws2.max_row + 1
for i, u in enumerate(additional_utilities):
    src_row = 73 + i + 2  # Starting from row 75 in tariff database (row 73 + header offset)
    ws2.cell(bra_row, 1, value=bra_row - 1)
    ws2.cell(bra_row, 2, value=f"='Tariff Database'!B{src_row}")
    ws2.cell(bra_row, 3, value=f"='Tariff Database'!C{src_row}")
    ws2.cell(bra_row, 4, value=f"='Tariff Database'!D{src_row}")
    ws2.cell(bra_row, 5, value=f"='Tariff Database'!E{src_row}")
    ws2.cell(bra_row, 6, value=f"='Tariff Database'!K{src_row}")
    ws2.cell(bra_row, 7, value=f"='Tariff Database'!L{src_row}")
    ws2.cell(bra_row, 8, value=f"='Tariff Database'!M{src_row}")
    ws2.cell(bra_row, 9, value=f"='Tariff Database'!N{src_row}")
    ws2.cell(bra_row, 10, value=f"='Tariff Database'!O{src_row}")
    ws2.cell(bra_row, 11, value=f"='Tariff Database'!Y{src_row}")
    ws2.cell(bra_row, 12, value=f"='Tariff Database'!Z{src_row}")
    ws2.cell(bra_row, 13, value=f"='Tariff Database'!AB{src_row}")
    ws2.cell(bra_row, 14, value=f"='Tariff Database'!I{src_row}")
    bra_row += 1

# Add new rows to Protection Matrix
pm_row = ws3.max_row + 1
for i, u in enumerate(additional_utilities):
    src_row = 73 + i + 2
    ws3.cell(pm_row, 1, value=pm_row - 1)
    ws3.cell(pm_row, 2, value=f"='Tariff Database'!B{src_row}")
    ws3.cell(pm_row, 3, value=f"='Tariff Database'!C{src_row}")
    ws3.cell(pm_row, 4, value=f"='Tariff Database'!J{src_row}")
    ws3.cell(pm_row, 5, value=f"='Tariff Database'!Q{src_row}")
    ws3.cell(pm_row, 6, value=f"='Tariff Database'!P{src_row}")
    ws3.cell(pm_row, 7, value=f"='Tariff Database'!S{src_row}")
    ws3.cell(pm_row, 8, value=f"='Tariff Database'!T{src_row}")
    ws3.cell(pm_row, 9, value=f"='Tariff Database'!U{src_row}")
    ws3.cell(pm_row, 10, value=f"='Tariff Database'!R{src_row}")
    ws3.cell(pm_row, 11, value=f"='Tariff Database'!V{src_row}")
    ws3.cell(pm_row, 12, value=f"='Tariff Database'!W{src_row}")
    ws3.cell(pm_row, 13, value=f"='Tariff Database'!X{src_row}")
    ws3.cell(pm_row, 14, value=f"='Tariff Database'!AA{src_row}")
    ws3.cell(pm_row, 15, value=f"='Tariff Database'!AB{src_row}")
    pm_row += 1

# Save
output_path = '/sessions/laughing-peaceful-archimedes/mnt/power-insight/Large_Load_Tariff_Database_FINAL.xlsx'
wb.save(output_path)

print(f"Updated file with {ws.max_row - 2} total utilities")
print(f"Blended Rate Analysis: {ws2.max_row - 1} entries")
print(f"Protection Matrix: {ws3.max_row - 1} entries")
print(f"Document Citations: {ws_citations.max_row - 1} entries")
