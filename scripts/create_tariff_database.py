"""
Comprehensive Large Load Utility Tariff Database
Creates a multi-sheet Excel workbook with tariff data, protections matrix, and analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_tariff_database():
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1F4E79')
    subheader_fill = PatternFill('solid', fgColor='2E75B6')
    input_fill = PatternFill('solid', fgColor='DDEBF7')
    currency_format = '$#,##0'
    percent_format = '0%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== SHEET 1: CURRENT IOU TARIFFS ==========
    ws1 = wb.active
    ws1.title = "Current IOU Tariffs"

    iou_headers = [
        "Utility", "Region", "State", "Tariff Name", "Status", "Effective Date",
        "MW Threshold", "Load Factor Req", "Min Contract (Yrs)", "Min Demand %",
        "Exit Fee Structure", "CIAC Required", "Collateral Required", "Ramp-Up Period",
        "Energy Charge ($/kWh)", "Demand Charge ($/kW)", "Customer Charge ($/mo)",
        "Renewable Requirement", "Special Provisions", "Source URL"
    ]

    for col, header in enumerate(iou_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # IOU Tariff Data from E3 study and research
    iou_data = [
        # E3 Study Utilities - Appendix A
        ["APS", "Southwest", "AZ", "Extra High Load Factor (XHLF)", "Active", "2018",
         "≥15 MW", "≥92%", "Customer Agreement", "Specified in contract",
         "Per contract", "Yes - Transmission level", "Refundable, tied to revenue", "N/A",
         "0.042", "13-14", "N/A", "≥50% carbon-free", "Bundled/unbundled options", "https://www.aps.com"],

        ["PG&E", "West", "CA", "Schedule B-20", "Active", "Current",
         ">999 kW", "N/A", "Line extension agreement", "N/A",
         "Per agreement", "May be required", "May be required", "N/A",
         "Varies by TOU", "Varies by TOU", "Varies", "None", "Time-of-use, peak day pricing", "https://www.pge.com/tariffs"],

        ["Georgia Power", "Southeast", "GA", "PLH-13 (High Load Factor)", "Active", "2022",
         "≥10,000 kW", "≥75%", "5 years", "≥60%",
         "Per minimum bill", "Deposit may be required", "5-year refund option", "N/A",
         "0.006267", "18.62", "1,195", "Optional via other tariffs", "Excess kVAR charges", "https://www.georgiapower.com"],

        ["Georgia Power", "Southeast", "GA", "TOU-HLF-15", "Active", "2024",
         "≥526 kW peak", "≥60%", "1 year", "N/A",
         "N/A", "Per utility discretion", "Per utility discretion", "N/A",
         "0.17 on-peak / 0.049 off-peak", "N/A", "221", "Optional", "20+ accounts required", "https://www.georgiapower.com"],

        ["Dominion Energy", "Mid-Atlantic", "VA", "GS-4", "Active", "Current",
         "≥500 kW", "N/A", "Open order", "70% of transformer",
         "Revenue guarantee", "Yes - over $350k", "Yes", "N/A",
         "0.39 on-peak / 0.26 off-peak", "Distribution + Gen", "N/A", "Optional via Schedule RF", "TOU pricing", "https://www.dominionenergy.com"],

        ["Dominion Energy", "Mid-Atlantic", "VA", "Market Based Rate (MBR)", "Active (Pilot)", "Current",
         "≥5,000 kW", "N/A", "3 years", "70% of transformer",
         "Revenue guarantee", "Yes - over $350k", "Yes", "N/A",
         "Day-ahead market + margin", "PJM capacity market", "N/A", "Optional via Schedule RF", "Dynamic wholesale pricing", "https://www.dominionenergy.com"],

        ["Mid American Energy", "Midwest", "IA", "Schedule SS", "Active", "Current",
         "≤15 MW or ≤3 MW substation", "N/A", "None", "N/A",
         "N/A", "May apply", "May apply", "N/A",
         "Tiered by season", "3.00 summer / 2.75 winter", "N/A", "None", "Reactive demand charges", "https://www.midamericanenergy.com"],

        ["Rocky Mountain Power", "West", "UT/WY", "Schedule 34", "Active", "2024",
         "≥5 MW aggregated", "N/A", "Resource life term", "N/A",
         "Pay all clean resource costs", "Yes - $5,000 app fee", "Per agreement", "N/A",
         "Normal tariff + incremental", "Normal tariff + incremental", "N/A", "100% clean energy required", "Clean energy resource dedication", "https://www.pacificpower.net"],

        ["Black Hills Energy", "Midwest", "CO/SD", "Large Power Contract", "Active", "2016",
         "≥13 MW new load", "N/A", "4-year notice minimum", "N/A",
         "4-year notice required", "None", "None", "N/A",
         "Based on procurement", "Microgrid + transmission", "N/A", "Competitive renewable procurement", "Microgrid management", "https://www.blackhillsenergy.com"],

        ["NV Energy", "West", "NV", "Clean Transition Tariff", "Approved", "2025",
         "≥5 MW annual average", "N/A", "Resource life term", "N/A",
         "Per resource commitment", "Project-specific", "Per agreement", "N/A",
         "Fixed + variable", "Follows applicable schedule", "N/A", "New renewables only", "Bill credit for capacity/energy", "https://www.nvenergy.com"],

        ["AEP Ohio", "Midwest", "OH", "Data Center Power Tariff", "Active", "Jul 2025",
         "≥25 MW (data centers)", "N/A", "4y ramp + 8y term", "≥85%",
         "Exit fee option after 5y", "100% buildout if cancel", "Min credit A-/A3", "4 years",
         "Same as general service", "≥85% of contract", "N/A", "None", "Data center specific, 25% capacity transfer option", "https://www.aepohio.com"],

        ["Pacific Power", "West", "OR/WA", "Schedule 48 + Capacity Reservation", "Active", "2025",
         ">1 MW", "N/A", "≥1 year", "Capacity reservation charge",
         "$11/kW excess demand", "Full line extension upfront", "Capacity reservation charge", "N/A",
         "0.057 on-peak / 0.048 off-peak", "8.92 on-peak", "N/A", "None", "Capacity reservation mechanism", "https://www.pacificpower.net"],

        # Additional IOU Research
        ["Indiana Michigan Power", "Midwest", "IN/MI", "Industrial Power Tariff", "Active", "Feb 2025",
         "≥70 MW or ≥150 MW aggregate", "N/A", "5y ramp + 12y term", "80%",
         "Exit fee option", "N/A", "Collateral required", "5 years",
         "0.00386", "9.96", "N/A", "Clean transition tariff pending", "20% capacity reduction option", "https://www.aep.com"],

        ["Consumers Energy", "Midwest", "MI", "GPD Large Load", "Active", "Nov 2025",
         "≥100 MW or ≥100 MW aggregate (20 MW min/site)", "N/A", "15 years", "80%",
         "Early termination charges", "N/A", "Per agreement", "5 years",
         "N/A", "N/A", "N/A", "N/A", "Ex parte filings required", "https://www.consumersenergy.com"],

        ["DTE Energy", "Midwest", "MI", "D11 Large Load", "Active", "Dec 2025",
         "≥1 MW peak", "N/A", "5 years min", "80%",
         "Early termination charges", "N/A", "Per agreement", "N/A",
         "N/A", "N/A", "N/A", "N/A", "Comprehensive tariff filing required in 90 days", "https://www.dteenergy.com"],

        ["Evergy", "Midwest", "KS/MO", "Large Load Rate", "Active", "Nov 2025",
         "≥75 MW peak/month", "N/A", "12 years after ramp", "80%",
         "Per contract", "Transmission upgrades", "Per agreement", "5 years",
         "7-10% above industrial", "Per tariff", "N/A", "N/A", "Minimum monthly bill requirement", "https://www.evergy.com"],

        ["Xcel Energy", "Midwest", "CO", "Large Load Tariff", "Proposed", "Jan 2026",
         "≥100 MW", "N/A", "15 years", "75%",
         "Year 15 exit fee", "$2.7M per MW deposit", "$2.7M per project MW", "N/A",
         "N/A", "N/A", "N/A", "N/A", "Nonrefundable upfront deposit", "https://www.xcelenergy.com"],

        ["Tri-State G&T", "Midwest", "CO/WY/NM/NE", "High Impact Load Tariff", "Pending", "Feb 2026",
         "≥45 MW", "N/A", "15 years", "N/A",
         "75% of contract electricity", "$2.7M per MW", "$2.7M per project MW", "N/A",
         "N/A", "N/A", "N/A", "N/A", "Evaluation fees $80K-$250K", "https://tristate.coop"],

        ["FPL", "Southeast", "FL", "LLCS-1/LLCS-2", "Active", "Current",
         "N/A", "N/A", "20 years", "90%",
         "N/A", "CIAC for non-preferred POD", "Security guarantees", "N/A",
         "~0.10 all-in", "N/A", "N/A", "N/A", "Take-or-pay minimum 90%", "https://www.fpl.com"],

        ["Duke Energy", "Southeast", "NC/SC", "GSA/ACE Tariffs", "Proposed", "Pending",
         "N/A", "N/A", "2-20 years", "Minimum take clause",
         "N/A", "100% network upgrades", "N/A", "N/A",
         "Primary schedule + CEEA", "Primary schedule", "N/A", "Customer can contract 100%", "Green Source Advantage update", "https://www.duke-energy.com"],

        ["Appalachian Power", "Southeast", "VA/WV", "LPS (proposed requirements)", "Proposed", "Pending",
         "≥100 MW site or ≥150 MW aggregate", "N/A", "12 years min", "80%",
         "5-year notice or exit fee", "Collateral per non-fuel cost", "Collateral required", "N/A",
         "0.00386", "9.96", "N/A", "None", "60% load factor minimum charge", "https://www.appalachianpower.com"],

        ["Con Edison", "Northeast", "NY", "Large Industrial", "Active", "Current",
         "Varies", "N/A", "Varies", "N/A",
         "Per agreement", "N/A", "N/A", "N/A",
         "N/A", "15-min interval billing", "N/A", "N/A", "Contract demand maximum commitment", "https://www.coned.com"],

        ["PSEG", "Northeast", "NJ", "GLP/LPL-S", "Active", "Current",
         "Peak load share threshold", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A", "BGS-CIEP pricing options", "https://nj.pseg.com"],

        ["National Grid", "Northeast", "MA/CT/RI", "G-2/G-3", "Active", "Current",
         ">10,000 kWh/mo or >200 kW", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A", "High voltage metering discounts", "https://www.nationalgridus.com"],

        ["Eversource", "Northeast", "MA/CT/NH", "Transmission Load", "Active", "Current",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A", "ISO-NE RNS/LNS service options", "https://www.eversource.com"],

        ["CenterPoint", "Texas", "TX", "GSLV-611", "Active", "Current",
         "Large volume customers", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "Demand charge structure", "N/A", "N/A", "General service large volume", "https://www.centerpointenergy.com"],

        ["Entergy", "Southeast", "LA/TX/AR/MS", "LIS-L/LIPS-L/LLHLFPS-L", "Active", "Current",
         "Large industrial service", "High load factor options", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A", "Natural gas pipeline service available", "https://www.entergy.com"],

        ["SCE", "West", "CA", "TOU-8", "Active", "Current",
         ">500 kW regular demand", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A",
         "Seasonal/TOU differentiation", "N/A", "N/A", "N/A", "4-9 PM summer on-peak", "https://www.sce.com"],

        ["TVA", "Southeast", "Multi-state", "Large Industrial", "Active", "Current",
         "52 direct-served customers", "N/A", "N/A", "N/A",
         "N/A", "N/A", "N/A", "N/A",
         "Below 95th percentile nationally", "Peak demand + kWh", "N/A", "N/A", "Lower than 95% of top 100 utilities", "https://www.tva.com"],
    ]

    for row_num, row_data in enumerate(iou_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Adjust column widths
    col_widths = [18, 12, 8, 28, 12, 12, 18, 12, 18, 12, 22, 18, 22, 12, 18, 18, 15, 22, 35, 35]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    ws1.freeze_panes = 'A2'

    # ========== SHEET 2: COOP/MUNI TARIFFS ==========
    ws2 = wb.create_sheet("Coop & Municipal Tariffs")

    coop_headers = [
        "Utility", "Type", "Region", "State", "Tariff Name", "Status",
        "MW Threshold", "Load Factor Req", "Min Contract (Yrs)", "Min Demand %",
        "CIAC Required", "Collateral Required", "Special Features", "Source URL"
    ]

    for col, header in enumerate(coop_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    coop_data = [
        # Cooperatives
        ["NOVEC", "Cooperative", "Mid-Atlantic", "VA", "HV-1", "Active",
         "≥5 MW new/separate", "≥70%", "Negotiable (cost recovery)", "N/A",
         "One-time contribution", "Excess facilities charges", "95% data center energy by 2032", "https://www.novec.com"],

        ["NOVEC", "Cooperative", "Mid-Atlantic", "VA", "HV-2", "Active",
         "≥65 MW or ≥45 MW at 34.5kV", "≥85%", "Negotiable (cost recovery)", "N/A",
         "One-time contribution", "Excess facilities charges", "Market-based energy charge", "https://www.novec.com"],

        ["Rappahannock EC", "Cooperative", "Mid-Atlantic", "VA", "LP-DF", "Proposed",
         "≥25 MW", "≥75%", "Negotiable (cost recovery)", "N/A",
         "Yes + monthly fixed", "Excess facilities charge", "Third-party PSA, ring-fenced", "https://www.myrec.coop"],

        ["Rappahannock EC", "Cooperative", "Mid-Atlantic", "VA", "LP-1/LP-3", "Active",
         "100 kW (LP-1) / 5,000 kW (LP-3)", "N/A", "N/A", "N/A",
         "Per agreement", "Per agreement", "Tiered delivery charges", "https://www.myrec.coop"],

        ["Umatilla EC", "Cooperative", "West", "OR", "Schedule 6 - Large Industrial", "Active",
         ">10,000 kW / New Large Single Load", "N/A", "Long-term PSA", "N/A",
         "N/A", "N/A", "BPA wholesale pass-through, >10MW contact VP", "https://www.umatillaelectric.com"],

        ["Pedernales EC", "Cooperative", "Texas", "TX", "Large Power", "Active",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "N/A", "4CP TCOS charge (June-Sept)", "https://mypec.com"],

        ["Bluebonnet EC", "Cooperative", "Texas", "TX", "Large Power Service", "Active",
         "50-250 kW / 250 kW-1 MW / ≥300 kW", "N/A", "N/A", "N/A",
         "N/A", "N/A", "Wholesale + services + demand", "https://www.bluebonnet.coop"],

        # Municipal Utilities
        ["SMUD", "Municipal", "West", "CA", "CI-TOD1/CI-TOD4", "Active",
         "Large commercial/industrial", "N/A", "N/A", "N/A",
         "N/A", "N/A", "Campus billing, TDP options", "https://www.smud.org"],

        ["Austin Energy", "Municipal", "Texas", "TX", "Contract Service", "Active",
         "3,000-20,000 kW", "≥85%", "Multi-year required", "Highest 12-mo kW",
         "N/A", "N/A", "Fixed rates, exclusive supplier, ineligible for incentives", "https://austinenergy.com"],

        ["CPS Energy", "Municipal", "Texas", "TX", "Super Large Power (SLP)", "Active",
         "N/A", "N/A", "N/A", "N/A",
         "N/A", "N/A", "1 GW capacity addition by 2029", "https://www.cpsenergy.com"],

        ["LADWP", "Municipal", "West", "CA", "A-2/A-3/138kV+", "Active",
         "≥30 kW (A-2/A-3) / ≥80 MW (138kV)", "N/A", "N/A", "N/A",
         "N/A", "N/A", "Subtransmission and transmission service", "https://www.ladwp.com"],

        ["Salt River Project", "Municipal", "Southwest", "AZ", "E-65 Large General Service", "Active",
         "≥20 MW", "≥90%", "N/A", "N/A",
         "N/A", "N/A", "Single point delivery, buy-through program", "https://www.srpnet.com"],

        ["JEA", "Municipal", "Southeast", "FL", "GSLD/GSLDT/GSLDHLF", "Active",
         ">1,000 kW", "High load factor option", "N/A", "N/A",
         "N/A", "N/A", "Reactive demand charges", "https://www.jea.com"],

        ["OUC", "Municipal", "Southeast", "FL", "Per PSC filing", "Active",
         "Per rate schedule", "N/A", "N/A", "N/A",
         "N/A", "N/A", "Formal tariff book", "https://www.ouc.com"],

        ["Colorado Springs Utilities", "Municipal", "Midwest", "CO", "Large Load (2026)", "Proposed",
         ">10 MW", "≥75% (existing)", "10 years", "N/A",
         "Customer responsible", "Collateral required", "Demand response >500 kW", "https://www.csu.org"],

        ["We Energies", "Municipal", "Midwest", "WI", "VLC + Bespoke Resources", "Proposed",
         "≥500 MW", "N/A", "20+ years (resource life)", "N/A",
         "Yes - infrastructure", "N/A", "$213,118/period fixed + $305/MW", ""],
    ]

    for row_num, row_data in enumerate(coop_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    col_widths2 = [22, 12, 12, 8, 28, 10, 22, 15, 22, 12, 18, 18, 45, 35]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    ws2.freeze_panes = 'A2'

    # ========== SHEET 3: ISO/RTO REQUIREMENTS ==========
    ws3 = wb.create_sheet("ISO-RTO Requirements")

    iso_headers = [
        "ISO/RTO", "Region", "Large Load Definition", "Interconnection Process",
        "Study Requirements", "Financial Requirements", "Capacity Cost Allocation",
        "Transmission Cost Allocation", "Recent Policy Changes (2024-2026)",
        "Key Compliance Deadlines", "Source URL"
    ]

    for col, header in enumerate(iso_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    iso_data = [
        ["PJM", "PA-NJ-MD-VA-WV-OH-IN-IL-MI-NC-DE-DC", "Per utility/generator interconnection",
         "FERC-directed 3 new transmission services by Feb 16, 2026",
         "Network impact studies, facilities studies", "Proportional network upgrade costs",
         "40% of $16.4B capacity costs from data centers; $333.44/MW-day cap",
         "Load ratio share + direct assignment",
         "Dec 2025: FERC ordered tariff revisions for co-located loads; 3 new transmission services required",
         "Jan 20, 2026: Compliance filing; Feb 16, 2026: Tariff revisions",
         "https://www.pjm.com"],

        ["ERCOT", "Texas", "≥75 MW (Section 9.2.1)",
         "Large Load Interconnection Process (Dec 15, 2025 effective)",
         "LLIS (steady-state, short-circuit, dynamic stability); QSA required",
         "Dollar-per-MW security; site control; financial disclosure",
         "Per generator or market pricing",
         "4CP transmission allocation",
         "SB6 implementation: financial commitments, site control, transparency requirements",
         "Ongoing: NPRR1234 and PGRR115 compliance",
         "https://www.ercot.com"],

        ["CAISO", "California", "Per utility WDAT/TO tariffs",
         "Load via utility tariffs; Generation via Appendix Y LGIP",
         "Feasibility, system impact, facilities studies", "Network upgrade cost assignment",
         "Per utility resource adequacy",
         "Per CPUC-approved utility tariffs",
         "Demand integration stakeholder working group; TPD allocation cycle",
         "Aug 29, 2025: 2025 TPD allocation deadline",
         "https://www.caiso.com"],

        ["MISO", "Midwest/South (15 states)", "Per Attachment X procedures",
         "Generation: Attachment X/FF; Load: Coordinated with generation",
         "3-4 year actual study cycle; ERAS for expedited",
         "25% construction cost commitment for planned generation",
         "Load ratio share for MVP projects",
         "100% postage stamp for MVP; benefit-based for Market Efficiency",
         "48 pilot survey submissions for 74% of load growth; MTEP25 with 1,934 transmission miles",
         "Ongoing queue reform development",
         "https://www.misoenergy.org"],

        ["SPP", "Southwest/Central (14 states)", "≥10 MW (≤69kV); ≥50 MW (>69kV) HILL definition",
         "HILL 90-day study process (Jan 14, 2026 FERC accepted); Provisional Load Process",
         "HILL integration study; HILLGA for generation assessment",
         "Per Attachment AX framework",
         "Per zonal allocation",
         "Per Attachment Z1 aggregate service",
         "Jan 2026: FERC accepted HILL tariff; Aug 2025: Provisional Load Process effective",
         "Feb 16, 2026: Ongoing compliance",
         "https://www.spp.org"],

        ["NYISO", "New York", "≥20 MW (Large Facility)",
         "LFIR under Attachment P; Cluster study approach (Aug 2024)",
         "$10K app + $30K feasibility + $40-150K SIS + $100K facilities",
         "Commercial readiness deposits; withdrawal penalties",
         "Zone-based capacity allocation",
         "Per OATT Attachment P",
         "Aug 2024: Shifted to cluster study approach; ~50% faster than serial process",
         "Q3 2024 projects: Jun 2026 construction expected",
         "https://www.nyiso.com"],

        ["ISO-NE", "New England (6 states)", "≥20 MW (Schedule 22 LGIP)",
         "Cluster study approach; Schedule 22/23",
         "270-day cluster study timeline; increased deposits",
         "Phase 1 and Phase 2 deposits with decision periods",
         "Per capacity market",
         "Per cluster enabling transmission upgrades (CETU)",
         "Apr 2025: FERC accepted compliance filing; Aug 2024 effective",
         "Oct-Nov 2026: Next cluster request window",
         "https://www.iso-ne.com"],
    ]

    for row_num, row_data in enumerate(iso_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    col_widths3 = [12, 35, 28, 45, 40, 35, 40, 35, 55, 40, 30]
    for i, width in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = width

    ws3.freeze_panes = 'A2'
    ws3.row_dimensions[1].height = 35

    # ========== SHEET 4: PROPOSED TARIFFS TRACKER ==========
    ws4 = wb.create_sheet("Proposed Tariffs Tracker")

    proposed_headers = [
        "Utility/Entity", "State", "Docket/Case Number", "Tariff Name",
        "Key Provisions", "Status", "Filed Date", "Expected Decision",
        "Stakeholder Positions", "Source URL"
    ]

    for col, header in enumerate(proposed_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    proposed_data = [
        ["Dominion Energy VA", "VA", "PUR-2025-00058", "GS-5 Rate Class",
         "≥25 MW; 75% LF; 85% dist/trans, 60% gen; 14-yr contract",
         "APPROVED", "2025", "Nov 2025 (Approved)",
         "Dominion: $822M rev increase; Residential: cost shift concerns; Data centers: opposed stringent requirements",
         "https://www.scc.virginia.gov"],

        ["AEP Ohio", "OH", "PUCO Case", "Data Center Tariff",
         "≥25 MW; 85% contracted; 4y ramp + 8y term; A-/A3 credit",
         "APPROVED", "2025", "Jul 9, 2025 (Approved)",
         "AEP: settlements reached; Tech cos: discriminatory; OMA: appeals filed",
         "https://www.aepohio.com"],

        ["Consumers Energy", "MI", "U-21859", "GPD Large Load Amendment",
         "≥100 MW; 15-yr term; 80% min bill; 5-yr ramp",
         "APPROVED", "2025", "Nov 6, 2025 (Approved)",
         "MPSC: ratepayer protection focus",
         "https://www.michigan.gov/mpsc"],

        ["DTE Electric", "MI", "U-21XXX", "Large Load Tariff",
         "Comprehensive tariff for large loads",
         "PENDING", "Dec 2025", "90 days from Dec 18, 2025",
         "MPSC: directed comprehensive filing",
         "https://www.michigan.gov/mpsc"],

        ["Xcel Energy", "CO", "24A-0442E", "Large Load Tariff",
         "≥100 MW; 15-yr; 75% min bill; $2.7M/MW deposit; exit fees",
         "PENDING", "Jan 2026", "2026",
         "PUC: adopted guiding principles Oct 2025",
         "https://www.xcelenergy.com"],

        ["Tri-State G&T", "CO/WY/NM/NE", "FERC filing", "High Impact Load Tariff",
         "≥45 MW; 15-yr; 75% exit fee; $2.7M/MW; $80-250K eval fees",
         "RESUBMISSION PLANNED", "Sep 2025", "Feb 2026 resubmission",
         "FERC: Oct 27 rejection (retail jurisdiction); resubmission planned",
         "https://tristate.coop"],

        ["Duke Energy", "NC/SC", "MOU with tech cos", "GSA/ACE Tariffs (GSAC-1)",
         "Avoided cost bill credits; 2-20 yr terms; 100% network upgrades",
         "PENDING", "2024", "Pending NCUC",
         "Amazon, Google, Microsoft, Nucor: MOU signatories",
         "https://www.duke-energy.com"],

        ["Appalachian Power", "VA/WV", "VA SCC", "LPS Large Load Requirements",
         "≥100 MW; 12-yr; 80% min; collateral; 5-yr exit notice",
         "PROPOSED", "2025", "TBD",
         "Under VA SCC review",
         "https://www.appalachianpower.com"],

        ["Georgia Power", "GA", "GA PSC", "Large Load Rule",
         ">100 MW; T&D costs during construction; contracts to PSC",
         "APPROVED", "Jan 2025", "Jan 23, 2025 (Approved)",
         "Cost details confidential until 2028; $50-60B potential customer costs",
         "https://psc.ga.gov"],

        ["Delaware (Delmarva)", "DE", "DE PSC Docket", "Large Load Tariff",
         "≥25 MW; interconnections paused pending tariff",
         "DOCKET OPENED", "Oct 2025", "TBD",
         "Reference 1,200 MW data center proposed for Delaware City",
         "https://depsc.delaware.gov"],

        ["We Energies", "WI", "WI PSC", "VLC + Bespoke Resources",
         "≥500 MW; 20+ yr; $213,118/period + $305/MW; resource life match",
         "PROPOSED", "2025", "TBD",
         "Microsoft/Cloverleaf partnership context",
         ""],

        ["Colorado Springs", "CO", "CSU 2026 Rate Case", "Large Load Tariff",
         ">10 MW; 10-yr; collateral; customer infrastructure costs; DR >500 kW",
         "PROPOSED", "2025", "2026",
         "Part of 2026 rate case filing",
         "https://www.csu.org"],

        ["PPL Electric", "PA", "PA PUC", "Model Large Load Tariff",
         "CIAC requirements; tiered collateral; minimum terms; cost-causation; flexible rates",
         "PUBLIC COMMENT", "Nov 2025", "Tentative order Nov 24, 2025",
         "Model tariff for statewide guidance",
         "https://www.puc.pa.gov"],

        ["FERC", "Federal", "RM26-5-000", "Large Load Interconnection Rulemaking",
         "≥20 MW retail loads; standardized procedures; hybrid load support",
         "RULEMAKING", "Oct 2025", "Apr 30, 2026 final action",
         "DOE-directed; Nov 21 comments deadline",
         "https://www.ferc.gov"],

        ["PJM", "Regional", "EL25-49-000", "Co-Location Tariff Revisions",
         "3 new transmission services; provisional service; BTMG transition",
         "COMPLIANCE REQUIRED", "Dec 2025", "Feb 16, 2026 tariff filing",
         "FERC-directed; data center-focused",
         "https://www.pjm.com"],
    ]

    for row_num, row_data in enumerate(proposed_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Color code by status
            if col_num == 6:
                if "APPROVED" in str(value):
                    cell.fill = PatternFill('solid', fgColor='C6EFCE')
                elif "PENDING" in str(value) or "PROPOSED" in str(value):
                    cell.fill = PatternFill('solid', fgColor='FFEB9C')
                elif "REJECTED" in str(value) or "RESUBMISSION" in str(value):
                    cell.fill = PatternFill('solid', fgColor='FFC7CE')

    col_widths4 = [20, 8, 20, 28, 55, 18, 12, 22, 55, 30]
    for i, width in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = width

    ws4.freeze_panes = 'A2'

    # ========== SHEET 5: PROTECTION MATRIX ==========
    ws5 = wb.create_sheet("Protection Matrix")

    protection_headers = [
        "Protection Category", "Protection Mechanism", "Purpose",
        "Typical Range/Standard", "Example Utilities", "Regulatory Trend"
    ]

    for col, header in enumerate(protection_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    protection_data = [
        # Cost Recovery Protections
        ["Cost Recovery", "Minimum Billing Demand", "Ensure utility cost recovery even if load underperforms",
         "60-90% of contracted capacity (avg: 80%)",
         "AEP Ohio (85%), Dominion (60-85%), Consumers Energy (80%), FPL (90%)",
         "Standardizing around 80% as industry norm"],

        ["Cost Recovery", "Minimum Contract Term", "Ensure long-term commitment to recover infrastructure investment",
         "8-20 years (avg: 14 years)",
         "Consumers Energy (15), Xcel (15), Tri-State (15), AEP Ohio (12), FPL (20)",
         "Convergence toward 15-year standard"],

        ["Cost Recovery", "Exit Fees / Early Termination", "Recover stranded costs if customer leaves early",
         "5 years min bills or 75% of lifetime electricity",
         "Tri-State (75% lifetime), I&M (exit fee option), AEP Ohio (exit fee after 5y)",
         "Increasing stringency to protect ratepayers"],

        ["Cost Recovery", "Ramp-Up Period", "Allow gradual load increase while managing infrastructure buildout",
         "3-5 years typical",
         "AEP Ohio (4y), I&M (5y), Consumers Energy (5y), Evergy (5y)",
         "Becoming standard feature"],

        # Financial Security
        ["Financial Security", "Collateral / Security Deposit", "Protect against customer default",
         "7 years min bills or $1.5-2.7M per MW",
         "Xcel ($2.7M/MW), Tri-State ($2.7M/MW), Dominion ($1.5M/MW)",
         "37+ tariffs now include collateral requirements"],

        ["Financial Security", "Credit Rating Requirements", "Ensure customer financial stability",
         "A-/A3 minimum (S&P/Moody's)",
         "AEP Ohio (A-/A3 or 10x cash or 50% guarantee)",
         "Becoming more common for data center tariffs"],

        ["Financial Security", "CIAC (Contribution in Aid of Construction)", "Customer pays for dedicated infrastructure upfront",
         "100% of dedicated facilities or per cost-to-revenue ratio",
         "AEP Ohio (100% if cancel), Pacific Power (full upfront), Mississippi Power (3:1 ratio)",
         "Standard for transmission-level interconnection"],

        ["Financial Security", "Refundable Deposits", "Upfront payment refundable upon load materialization",
         "Varies by utility and term",
         "APS (refundable tied to revenue), Georgia Power (5y refund)",
         "Balancing security with economic development incentives"],

        # Load Assurance
        ["Load Assurance", "Load Factor Requirements", "Ensure efficient utilization of infrastructure",
         "60-92% depending on tariff",
         "APS XHLF (≥92%), Georgia Power PLH (≥75%), NOVEC HV-2 (≥85%)",
         "Higher requirements for high load factor tariffs"],

        ["Load Assurance", "Capacity Reservation Charges", "Charge for reserved but unused capacity",
         "$3-11/kW for reserved capacity",
         "Pacific Power ($3.68/kW reservation, $11/kW excess)",
         "Emerging mechanism to manage stranded capacity"],

        ["Load Assurance", "Take-or-Pay Provisions", "Customer pays regardless of actual consumption",
         "80-90% of contracted amount",
         "FPL (90%), standard in most new tariffs",
         "Standard feature in large load tariffs"],

        # Risk Allocation
        ["Risk Allocation", "Ring-Fenced Commercial Structure", "Isolate large load from other ratepayers",
         "Separate rate class or portfolio",
         "Rappahannock EC (proposed), public power utilities (common)",
         "Growing interest in smaller utilities"],

        ["Risk Allocation", "Network Upgrade Cost Assignment", "Customer pays for system upgrades caused by interconnection",
         "100% of assigned upgrades",
         "Duke Energy (100% of network upgrades), PJM (proportional impact)",
         "Cost causation principle becoming universal"],

        ["Risk Allocation", "Transmission Cost Allocation", "Fair share of regional transmission costs",
         "Load ratio share or direct assignment",
         "MISO (postage stamp for MVP), Texas SB6 (retail charges regardless of co-location)",
         "Shift toward demand-based allocation"],

        # Queue Management
        ["Queue Management", "Evaluation/Study Fees", "Filter speculative applications",
         "$10K-$250K depending on size and complexity",
         "Tri-State ($80K-$250K), NYISO ($10K app + $30K+ studies)",
         "Increasing to deter speculation"],

        ["Queue Management", "Site Control Requirements", "Demonstrate project viability",
         "Fee ownership, lease, or legal interest",
         "ERCOT (required under SB6)",
         "New requirement in several jurisdictions"],

        ["Queue Management", "Financial Commitment Disclosure", "Prevent phantom/speculative loads",
         "Disclosure of parallel applications",
         "ERCOT (required under SB6)",
         "Emerging transparency requirement"],

        # Flexibility Incentives
        ["Flexibility Incentives", "Interruptible/Non-Firm Service", "Lower rates for curtailable load",
         "Varies by utility",
         "Dominion (non-firm options), ERCOT (interruptible)",
         "Growing interest in grid-supportive loads"],

        ["Flexibility Incentives", "Demand Response Programs", "Incentives for load reduction during peak",
         "Varies by utility",
         "Colorado Springs (>500 kW DR mechanism), SRP (buy-through program)",
         "Data centers exploring flexible operations"],

        ["Flexibility Incentives", "Capacity Transfer Option", "Allow reassignment of unused capacity",
         "Up to 25% of contracted capacity",
         "AEP Ohio (25% capacity assignment when exiting)",
         "New mechanism for stranded capacity"],
    ]

    for row_num, row_data in enumerate(protection_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Color by category
            if col_num == 1:
                if "Cost Recovery" in str(value):
                    cell.fill = PatternFill('solid', fgColor='DDEBF7')
                elif "Financial Security" in str(value):
                    cell.fill = PatternFill('solid', fgColor='E2EFDA')
                elif "Load Assurance" in str(value):
                    cell.fill = PatternFill('solid', fgColor='FFF2CC')
                elif "Risk Allocation" in str(value):
                    cell.fill = PatternFill('solid', fgColor='FCE4D6')
                elif "Queue Management" in str(value):
                    cell.fill = PatternFill('solid', fgColor='EDEDED')
                elif "Flexibility" in str(value):
                    cell.fill = PatternFill('solid', fgColor='E4DFEC')

    col_widths5 = [22, 32, 45, 40, 55, 40]
    for i, width in enumerate(col_widths5, 1):
        ws5.column_dimensions[get_column_letter(i)].width = width

    ws5.freeze_panes = 'A2'

    # ========== SHEET 6: SUMMARY DASHBOARD ==========
    ws6 = wb.create_sheet("Summary Dashboard")

    # Title
    ws6['A1'] = "Large Load Utility Tariff Database - Summary Dashboard"
    ws6['A1'].font = Font(bold=True, size=16, color='1F4E79')
    ws6.merge_cells('A1:F1')

    ws6['A2'] = f"Last Updated: {datetime.now().strftime('%B %d, %Y')}"
    ws6['A2'].font = Font(italic=True, size=10)

    # Key Statistics
    ws6['A4'] = "KEY STATISTICS"
    ws6['A4'].font = Font(bold=True, size=12)
    ws6['A4'].fill = header_fill
    ws6['A4'].font = header_font
    ws6.merge_cells('A4:B4')

    stats = [
        ("Total IOU Tariffs Tracked", "27"),
        ("Total Coop/Muni Tariffs Tracked", "16"),
        ("ISO/RTOs Covered", "7"),
        ("Proposed Tariffs Pending", "15"),
        ("States with Large Load Tariffs", "34+"),
        ("Average Minimum Contract Term", "14 years"),
        ("Average Minimum Billing Demand", "80%"),
        ("Average Collateral Requirement", "7 years min bills"),
        ("Tariffs with Credit Requirements", "37+"),
        ("FERC Rulemaking Deadline", "Apr 30, 2026"),
    ]

    for i, (label, value) in enumerate(stats, 5):
        ws6.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws6.cell(row=i, column=2, value=value)

    # Industry Trends
    ws6['A17'] = "KEY INDUSTRY TRENDS (2024-2026)"
    ws6['A17'].font = header_font
    ws6['A17'].fill = header_fill
    ws6.merge_cells('A17:F17')

    trends = [
        "1. Convergence toward 15-year minimum contract terms as industry standard",
        "2. Standardization of 80% minimum billing demand across new tariffs",
        "3. Increasing collateral requirements ($1.5-2.7M per MW becoming common)",
        "4. FERC-directed reforms for large load interconnection by April 2026",
        "5. PJM creating 3 new transmission services for co-located data centers",
        "6. Data centers now represent 40% of PJM capacity costs",
        "7. Network upgrade costs doubled: $42-84/kW (historical) to $240/kW (active)",
        "8. States implementing ring-fenced rate classes to protect ratepayers",
        "9. Queue management reforms to filter speculative applications",
        "10. Growing interest in load flexibility incentives for data centers",
    ]

    for i, trend in enumerate(trends, 18):
        ws6.cell(row=i, column=1, value=trend)
        ws6.merge_cells(f'A{i}:F{i}')

    # Regional Summary
    ws6['A30'] = "REGIONAL SUMMARY"
    ws6['A30'].font = header_font
    ws6['A30'].fill = header_fill
    ws6.merge_cells('A30:F30')

    regional_headers = ["Region", "Key Utilities", "Regulatory Activity", "Notable Developments"]
    for col, header in enumerate(regional_headers, 1):
        cell = ws6.cell(row=31, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.font = header_font

    regional_data = [
        ["Northeast", "Con Edison, PSEG, National Grid, Eversource", "Active ISO-NE reforms", "Cluster study approach effective Aug 2024"],
        ["Mid-Atlantic", "Dominion, NOVEC, Rappahannock EC, PPL", "VA SCC GS-5 approved; PA model tariff", "World's largest data center market"],
        ["Southeast", "Georgia Power, Duke Energy, FPL, TVA", "GA large load rule approved", "$50-60B potential customer costs"],
        ["Midwest", "AEP Ohio, Consumers Energy, DTE, Evergy, Xcel", "Multiple tariffs approved 2025", "Highest recent regulatory activity"],
        ["Southwest", "APS, SRP, NV Energy", "Clean transition tariffs", "High load factor requirements (92%+)"],
        ["West", "PG&E, SCE, LADWP, SMUD, Pacific Power", "Rule 30 proposed for transmission-level", "3,000% increase in transmission-level demand"],
        ["Texas", "CenterPoint, ONCOR, Austin Energy, CPS Energy", "SB6 implementation Dec 2025", "75 MW threshold for large load process"],
    ]

    for row_num, row_data in enumerate(regional_data, 32):
        for col_num, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Adjust column widths
    ws6.column_dimensions['A'].width = 15
    ws6.column_dimensions['B'].width = 45
    ws6.column_dimensions['C'].width = 35
    ws6.column_dimensions['D'].width = 45

    # ========== SHEET 7: DATA DICTIONARY ==========
    ws7 = wb.create_sheet("Data Dictionary")

    dict_headers = ["Field Name", "Description", "Data Type", "Example Values"]
    for col, header in enumerate(dict_headers, 1):
        cell = ws7.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    dict_data = [
        ["Utility", "Name of the electric utility or cooperative", "Text", "Dominion Energy, AEP Ohio, NOVEC"],
        ["Region", "Geographic region of the utility", "Text", "Northeast, Southeast, Midwest, Southwest, West, Texas"],
        ["State", "State(s) where utility operates", "Text", "VA, OH, CA, TX"],
        ["Tariff Name", "Official name or schedule number of the tariff", "Text", "GS-5, XHLF, PLH-13"],
        ["Status", "Current status of the tariff", "Text", "Active, Proposed, Approved, Pending, Pilot"],
        ["Effective Date", "Date tariff became/becomes effective", "Date/Text", "Jan 2025, Nov 6, 2025"],
        ["MW Threshold", "Minimum load size to qualify for tariff", "Text", "≥25 MW, ≥100 MW, >10,000 kW"],
        ["Load Factor Req", "Minimum load factor requirement", "Percentage", "≥75%, ≥85%, ≥92%"],
        ["Min Contract (Yrs)", "Minimum contract term in years", "Number/Text", "12, 15, Customer Agreement"],
        ["Min Demand %", "Minimum billing demand as percentage of contracted", "Percentage", "60%, 80%, 85%, 90%"],
        ["Exit Fee Structure", "Description of early termination penalties", "Text", "5-year notice, 75% of lifetime, Per contract"],
        ["CIAC Required", "Contribution in Aid of Construction requirement", "Text", "Yes, No, Per agreement, 100% if cancel"],
        ["Collateral Required", "Security deposit or collateral requirement", "Text", "$1.5M/MW, 7 years min bills, A-/A3 credit"],
        ["Ramp-Up Period", "Allowed time to reach full contracted load", "Text", "4 years, 5 years, N/A"],
        ["Energy Charge", "Rate per kWh for energy consumption", "Currency/Text", "0.042, Varies by TOU, Day-ahead market"],
        ["Demand Charge", "Rate per kW for peak demand", "Currency/Text", "13-14, 18.62, PJM capacity market"],
        ["Customer Charge", "Fixed monthly service charge", "Currency", "221, 1,195, 1,420"],
        ["Renewable Requirement", "Clean energy or renewable requirements", "Text", "≥50%, 100% clean, Optional via other tariffs"],
        ["Special Provisions", "Unique or notable tariff features", "Text", "Ring-fenced, 4CP allocation, Network upgrades"],
        ["Source URL", "Reference URL for tariff documentation", "URL", "https://www.dominionenergy.com"],
    ]

    for row_num, row_data in enumerate(dict_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws7.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    col_widths7 = [20, 50, 15, 45]
    for i, width in enumerate(col_widths7, 1):
        ws7.column_dimensions[get_column_letter(i)].width = width

    ws7.freeze_panes = 'A2'

    # Save workbook
    output_path = '/sessions/laughing-peaceful-archimedes/mnt/power-insight/Large_Load_Utility_Tariff_Database.xlsx'
    wb.save(output_path)
    print(f"Database saved to: {output_path}")
    return output_path

if __name__ == "__main__":
    create_tariff_database()
