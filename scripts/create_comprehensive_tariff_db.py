"""
Comprehensive Large Load Utility Tariff Database with Blended Rate Calculations

Assumptions for Blended Rate:
- Data Center Size: 600 MW
- Load Factor: 80%
- Average Load: 480 MW
- Monthly Hours: 730 (24 × 30.4 days)
- Monthly Consumption: 350,400 MWh = 350,400,000 kWh
- Peak/Off-Peak Split: 40% peak hours, 60% off-peak hours (conservative)
- Peak Consumption: 140,160,000 kWh
- Off-Peak Consumption: 210,240,000 kWh

Blended Rate = Total Monthly Cost / Total Monthly kWh
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Constants for blended rate calculation
DC_SIZE_MW = 600
LOAD_FACTOR = 0.80
AVG_LOAD_MW = DC_SIZE_MW * LOAD_FACTOR  # 480 MW
MONTHLY_HOURS = 730
MONTHLY_KWH = AVG_LOAD_MW * 1000 * MONTHLY_HOURS  # 350,400,000 kWh
PEAK_HOURS_PCT = 0.40
OFFPEAK_HOURS_PCT = 0.60
PEAK_KWH = MONTHLY_KWH * PEAK_HOURS_PCT  # 140,160,000 kWh
OFFPEAK_KWH = MONTHLY_KWH * OFFPEAK_HOURS_PCT  # 210,240,000 kWh
BILLING_DEMAND_KW = DC_SIZE_MW * 1000  # 600,000 kW

def calculate_blended_rate(tariff):
    """
    Calculate all-in blended rate for 600MW DC at 80% LF.
    Returns $/kWh blended rate.
    """
    # Get rates (default to 0 if not available)
    peak_demand = tariff.get('peak_demand_charge', 0) or 0
    offpeak_demand = tariff.get('off_peak_demand_charge', 0) or 0
    energy_peak = tariff.get('energy_rate_peak', 0) or 0
    energy_offpeak = tariff.get('energy_rate_off_peak', 0) or energy_peak * 0.7  # Estimate if not provided
    fixed_charge = tariff.get('fixed_charge', 500) or 500  # Default $500/month

    # Monthly demand cost
    # Use billing demand (may be ratcheted)
    min_demand_pct = tariff.get('min_demand_pct', 0) or 0
    if min_demand_pct > 0:
        billing_demand = max(BILLING_DEMAND_KW, BILLING_DEMAND_KW * (min_demand_pct / 100))
    else:
        billing_demand = BILLING_DEMAND_KW

    demand_cost = (peak_demand * billing_demand)
    if offpeak_demand > 0:
        demand_cost += (offpeak_demand * billing_demand * 0.5)  # Assume 50% off-peak demand charge applies

    # Monthly energy cost
    energy_cost = (energy_peak * PEAK_KWH) + (energy_offpeak * OFFPEAK_KWH)

    # Total monthly cost
    total_monthly = demand_cost + energy_cost + fixed_charge

    # Blended rate
    if MONTHLY_KWH > 0:
        blended_rate = total_monthly / MONTHLY_KWH
    else:
        blended_rate = 0

    return round(blended_rate, 5)

def calculate_protection_score(tariff):
    """Calculate protection score and return (score, category)."""
    score = 0

    min_demand = tariff.get('min_demand_pct', 0) or 0
    if min_demand >= 85: score += 3
    elif min_demand >= 75: score += 2
    elif min_demand >= 60: score += 1

    contract_years = tariff.get('contract_term_years', 0) or 0
    if contract_years >= 15: score += 3
    elif contract_years >= 10: score += 2
    elif contract_years >= 5: score += 1

    if tariff.get('ciac_required'): score += 2
    if tariff.get('take_or_pay'): score += 2
    if tariff.get('exit_fee'): score += 2
    if tariff.get('demand_ratchet'): score += 1
    if tariff.get('credit_requirements'): score += 1
    if tariff.get('dc_specific'): score += 2

    if score >= 12: return (score, 'High')
    elif score >= 7: return (score, 'Mid')
    else: return (score, 'Low')

# Comprehensive utility database
UTILITIES = [
    # ==================== OKLAHOMA ====================
    {'utility': 'Public Service Company of Oklahoma (PSO)', 'state': 'OK', 'region': 'Plains', 'iso_rto': 'SPP',
     'tariff_name': 'Large Power & Light (LPL)', 'rate_schedule': 'Schedule 242/244/246',
     'effective_date': '2025-01-30', 'status': 'Active', 'docket': 'PUD 2023-000086',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.05, 'off_peak_demand_charge': 2.47,
     'energy_rate_peak': 0.00171, 'energy_rate_off_peak': 0.00125,
     'contract_term_years': 7, 'min_demand_pct': 90, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'PSO Large Commercial and Industrial, 9th Revised Sheet No. 20',
     'source_url': 'https://www.psoklahoma.com/lib/docs/ratesandtariffs/Oklahoma/PSOLargeCommercialandIndustrialFeb2025.pdf',
     'page_reference': 'Sheet No. 20-3'},

    {'utility': 'Oklahoma Gas & Electric (OG&E)', 'state': 'OK', 'region': 'Plains', 'iso_rto': 'SPP',
     'tariff_name': 'Power & Light Large', 'rate_schedule': 'Schedule PL-1',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'PUD 2023-000087',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.08, 'off_peak_demand_charge': 2.50,
     'energy_rate_peak': 0.007, 'energy_rate_off_peak': 0.005,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'OG&E Oklahoma Rate Tariff, Schedule PL-1',
     'source_url': 'https://www.oge.com/documents/d/portal/19-00-oct-1-stamped-approved',
     'page_reference': 'Schedule 19.00'},

    {'utility': 'SWEPCO (AEP)', 'state': 'TX/LA/AR', 'region': 'Plains', 'iso_rto': 'SPP',
     'tariff_name': 'Large Load Contract', 'rate_schedule': 'ES-LL Contract',
     'effective_date': '2025-10-01', 'status': 'Active', 'docket': 'PUCT Filing Oct 2025',
     'min_load_mw': 75.0, 'peak_demand_charge': 8.20, 'off_peak_demand_charge': 3.10,
     'energy_rate_peak': 0.035, 'energy_rate_off_peak': 0.025,
     'contract_term_years': 12, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'SWEPCO Texas Tariff Manual',
     'source_url': 'https://www.swepco.com/lib/docs/ratesandtariffs/Texas/TexasRatesChargesandFees_08-28-25.pdf',
     'page_reference': 'ES-LL Section'},

    # ==================== TEXAS / ERCOT ====================
    {'utility': 'ERCOT (Grid Operator)', 'state': 'TX', 'region': 'Texas', 'iso_rto': 'ERCOT',
     'tariff_name': '4CP Transmission Allocation', 'rate_schedule': 'ERCOT Protocols',
     'effective_date': '2024-01-01', 'status': 'Active', 'docket': 'PUCT Project 52376',
     'min_load_mw': 1.0, 'peak_demand_charge': 5.50, 'off_peak_demand_charge': 1.50,
     'energy_rate_peak': 0.050, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'ERCOT Nodal Protocols, Section 7',
     'source_url': 'https://www.ercot.com/mktrules/nprotocols',
     'page_reference': 'Section 7.7 (4CP)'},

    {'utility': 'Oncor Electric Delivery', 'state': 'TX', 'region': 'Texas', 'iso_rto': 'ERCOT',
     'tariff_name': 'Large Load Delivery', 'rate_schedule': 'Rate Schedule 1700',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'PUCT Docket 54870',
     'min_load_mw': 1.0, 'peak_demand_charge': 3.80, 'off_peak_demand_charge': 1.20,
     'energy_rate_peak': 0.0285, 'energy_rate_off_peak': 0.020,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Oncor Tariff for Retail Delivery Service',
     'source_url': 'https://www.oncor.com/content/dam/oncorwww/documents/tariff/OnDTTariff.pdf',
     'page_reference': 'Schedule 1700'},

    {'utility': 'CenterPoint Energy Houston', 'state': 'TX', 'region': 'Texas', 'iso_rto': 'ERCOT',
     'tariff_name': 'General Service Large Volume', 'rate_schedule': 'GSLV-630',
     'effective_date': '2025-03-01', 'status': 'Active', 'docket': 'PUCT Docket 55678',
     'min_load_mw': 1.0, 'peak_demand_charge': 4.39, 'off_peak_demand_charge': 2.00,
     'energy_rate_peak': 0.04338, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'CenterPoint Energy Houston Electric Tariff',
     'source_url': 'https://www.centerpointenergy.com/en-us/corp/pages/rates-and-tariffs-electric.aspx',
     'page_reference': 'Schedule GSLV-630'},

    # ==================== SOUTHEAST ====================
    {'utility': 'Georgia Power', 'state': 'GA', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Power and Light Large', 'rate_schedule': 'Schedule PLL-11',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'GA PSC Docket 44280',
     'min_load_mw': 0.5, 'peak_demand_charge': 9.53, 'off_peak_demand_charge': 4.20,
     'energy_rate_peak': 0.0143, 'energy_rate_off_peak': 0.0095,
     'contract_term_years': 10, 'min_demand_pct': 95, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Georgia Power PLL-11 Rate Schedule',
     'source_url': 'https://www.georgiapower.com/content/dam/georgia-power/pdfs/business-pdfs/rates-schedules/small-business/PLL-11.pdf',
     'page_reference': 'PLL-11, Page 1'},

    {'utility': 'Duke Energy Carolinas', 'state': 'NC', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Schedule LGS',
     'effective_date': '2024-09-01', 'status': 'Active', 'docket': 'NC Docket E-7, Sub 1276',
     'min_load_mw': 1.0, 'peak_demand_charge': 5.20, 'off_peak_demand_charge': 3.50,
     'energy_rate_peak': 0.035, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 5, 'min_demand_pct': 70, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Duke Energy Carolinas Schedule LGS',
     'source_url': 'https://www.duke-energy.com/home/billing/rates',
     'page_reference': 'Schedule LGS'},

    {'utility': 'Duke Energy Florida', 'state': 'FL', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Large Load Customer', 'rate_schedule': 'Schedule LLC-1 (Proposed)',
     'effective_date': 'TBD', 'status': 'Proposed', 'docket': 'FL PSC Docket 20250113-EI',
     'min_load_mw': 50.0, 'peak_demand_charge': 7.73, 'off_peak_demand_charge': 2.71,
     'energy_rate_peak': 0.040, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 12, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Duke Energy Florida LLC-1 Filing, Document 09146-2025',
     'source_url': 'https://www.floridapsc.com/pscfiles/library/filings/2025/09146-2025/09146-2025.pdf',
     'page_reference': 'Docket 20250113-EI'},

    {'utility': 'Florida Power & Light (FPL)', 'state': 'FL', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Large Load Contract Service', 'rate_schedule': 'Schedule LLCS-1 (Proposed)',
     'effective_date': 'TBD', 'status': 'Proposed', 'docket': 'FL PSC Docket 20250011-EI',
     'min_load_mw': 50.0, 'peak_demand_charge': 8.50, 'off_peak_demand_charge': 3.20,
     'energy_rate_peak': 0.1016, 'energy_rate_off_peak': 0.065,
     'contract_term_years': 20, 'min_demand_pct': 90, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'FPL Electric Tariff Section 8',
     'source_url': 'https://www.fpl.com/rates/pdf/electric-tariff-section8.pdf',
     'page_reference': 'Section 8, LLCS-1'},

    {'utility': 'Alabama Power', 'state': 'AL', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Light and Power Service - Large', 'rate_schedule': 'Rate LPL',
     'effective_date': '2023-06-01', 'status': 'Active', 'docket': 'AL PSC Docket 24860',
     'min_load_mw': 1.0, 'peak_demand_charge': 12.50, 'off_peak_demand_charge': 4.80,
     'energy_rate_peak': 0.028, 'energy_rate_off_peak': 0.020,
     'contract_term_years': 5, 'min_demand_pct': 60, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Alabama Power LPL Rate Schedule',
     'source_url': 'https://www.alabamapower.com/content/dam/alabama-power/pdfs-docs/Rates/LPL.pdf',
     'page_reference': 'Rate LPL'},

    {'utility': 'Tampa Electric (TECO)', 'state': 'FL', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Large General Service Demand', 'rate_schedule': 'Schedule GSLD',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'FL PSC Docket 20240077',
     'min_load_mw': 0.5, 'peak_demand_charge': 8.20, 'off_peak_demand_charge': 3.40,
     'energy_rate_peak': 0.038, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'TECO Tariff Book Section 6',
     'source_url': 'https://www.tampaelectric.com/company/ourpowersystem/tariff/',
     'page_reference': 'Section 6, GSLD'},

    {'utility': 'TVA', 'state': 'TN/AL/KY/MS', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'General Service Rate', 'rate_schedule': 'GSA Part 3',
     'effective_date': '2024-10-01', 'status': 'Active', 'docket': 'TVA Board Approval',
     'min_load_mw': 1.0, 'peak_demand_charge': 5.34, 'off_peak_demand_charge': 2.50,
     'energy_rate_peak': 0.0245, 'energy_rate_off_peak': 0.0185,
     'contract_term_years': 5, 'min_demand_pct': 60, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'TVA GSA Rate Schedules',
     'source_url': 'https://www.tva.com/energy/valley-energy-rates',
     'page_reference': 'GSA Part 3'},

    {'utility': 'Santee Cooper', 'state': 'SC', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Large Light & Power', 'rate_schedule': 'Schedule LL&P-50MW',
     'effective_date': '2025-04-01', 'status': 'Active', 'docket': 'SC PSC Dec 2024 Approval',
     'min_load_mw': 50.0, 'peak_demand_charge': 11.20, 'off_peak_demand_charge': 4.50,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 15, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Santee Cooper Rate Study 2024',
     'source_url': 'https://www.santeecooper.com/Rates/Rate-Study/',
     'page_reference': 'Large Load Schedule'},

    {'utility': 'Dominion Energy SC', 'state': 'SC', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'Large Power Service', 'rate_schedule': 'Rate 27',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'SC PSC Docket 2024-34-E',
     'min_load_mw': 1.0, 'peak_demand_charge': 9.80, 'off_peak_demand_charge': 4.10,
     'energy_rate_peak': 0.042, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Dominion Energy SC Tariff Rate 27',
     'source_url': 'https://etariff.psc.sc.gov/Organization/Detail/411',
     'page_reference': 'Rate 27'},

    {'utility': 'JEA (Jacksonville)', 'state': 'FL', 'region': 'Southeast', 'iso_rto': 'None',
     'tariff_name': 'General Service Large Demand', 'rate_schedule': 'Schedule GSLD',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'JEA Board Approval',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.80, 'off_peak_demand_charge': 3.20,
     'energy_rate_peak': 0.042, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'JEA Electric Tariff',
     'source_url': 'https://www.jea.com/My_Account/Rates/Electric_Tariff.aspx',
     'page_reference': 'Schedule GSLD'},

    {'utility': 'Cleco', 'state': 'LA', 'region': 'Southeast', 'iso_rto': 'MISO',
     'tariff_name': 'Large Power Service', 'rate_schedule': 'Schedule LPS',
     'effective_date': '2024-07-01', 'status': 'Active', 'docket': 'LPSC Filing',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.50, 'off_peak_demand_charge': 3.00,
     'energy_rate_peak': 0.038, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Cleco Schedule LPS',
     'source_url': 'https://www.cleco.com/docs/default-source/rates-and-fees/rate-schedule/july-2024/07-lps-large-power-service-7-1-24.pdf',
     'page_reference': 'Schedule LPS'},

    {'utility': 'Entergy Louisiana', 'state': 'LA', 'region': 'Southeast', 'iso_rto': 'MISO',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Schedule LGS',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'LA PSC Filing',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.20, 'off_peak_demand_charge': 2.80,
     'energy_rate_peak': 0.042, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Entergy Louisiana ELL Tariffs',
     'source_url': 'https://www.entergylouisiana.com/business/ell-tariffs',
     'page_reference': 'Schedule LGS'},

    {'utility': 'Entergy Arkansas', 'state': 'AR', 'region': 'Southeast', 'iso_rto': 'MISO',
     'tariff_name': 'Large Industrial Service', 'rate_schedule': 'Schedule LIS-L',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'AR PSC Filing',
     'min_load_mw': 1.0, 'peak_demand_charge': 6.80, 'off_peak_demand_charge': 2.60,
     'energy_rate_peak': 0.038, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Entergy Arkansas Tariffs',
     'source_url': 'https://www.entergyarkansas.com/business/tariffs',
     'page_reference': 'Schedule LIS-L'},

    # ==================== PJM REGION ====================
    {'utility': 'AEP Ohio', 'state': 'OH', 'region': 'Midwest', 'iso_rto': 'PJM',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Schedule GS-4',
     'effective_date': '2024-06-01', 'status': 'Active', 'docket': 'PUCO Case 24-0XXX',
     'min_load_mw': 1.0, 'peak_demand_charge': 6.50, 'off_peak_demand_charge': 2.00,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 12, 'min_demand_pct': 85, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'AEP Ohio Tariff Book, Schedule GS-4',
     'source_url': 'https://www.aepohio.com/lib/docs/ratesandtariffs/ohio/aepohio-tariff.pdf',
     'page_reference': 'Schedule GS-4, Section III'},

    {'utility': 'Dominion Energy Virginia', 'state': 'VA', 'region': 'Mid-Atlantic', 'iso_rto': 'PJM',
     'tariff_name': 'Large General Service TOU', 'rate_schedule': 'Schedule GS-4',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'VA SCC Case PUR-2024-00067',
     'min_load_mw': 0.5, 'peak_demand_charge': 8.77, 'off_peak_demand_charge': 0.52,
     'energy_rate_peak': 0.027, 'energy_rate_off_peak': 0.018,
     'contract_term_years': 14, 'min_demand_pct': 85, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Dominion Virginia Electric Tariff',
     'source_url': 'https://www.dominionenergy.com/virginia/rates-and-tariffs',
     'page_reference': 'Schedule GS-4'},

    {'utility': 'NOVEC', 'state': 'VA', 'region': 'Mid-Atlantic', 'iso_rto': 'PJM',
     'tariff_name': 'Data Center Rate', 'rate_schedule': 'Schedule DC',
     'effective_date': '2024-01-01', 'status': 'Active', 'docket': 'VA SCC Case PUR-2023-00XXX',
     'min_load_mw': 5.0, 'peak_demand_charge': 12.50, 'off_peak_demand_charge': 6.80,
     'energy_rate_peak': 0.048, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 20, 'min_demand_pct': 90, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'NOVEC Schedule DC Data Center Rate',
     'source_url': 'https://www.novec.com/rates',
     'page_reference': 'Schedule DC'},

    {'utility': 'PPL Electric', 'state': 'PA', 'region': 'Mid-Atlantic', 'iso_rto': 'PJM',
     'tariff_name': 'Large C&I Transmission', 'rate_schedule': 'Pa. P.U.C. No. 201',
     'effective_date': '2025-06-01', 'status': 'Active', 'docket': 'PA PUC Docket M-2025-3054271',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.80, 'off_peak_demand_charge': 3.20,
     'energy_rate_peak': 0.042, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 5, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'PPL Electric Current Tariff Supplement 393-394',
     'source_url': 'https://www.pplelectric.com/utility/about-us/electric-rates-and-rules/current-electric-tariff.aspx',
     'page_reference': 'Supplement No. 393-394'},

    {'utility': 'PSEG', 'state': 'NJ', 'region': 'Mid-Atlantic', 'iso_rto': 'PJM',
     'tariff_name': 'Large Power & Lighting', 'rate_schedule': 'Schedule LPL-S/LPL-P',
     'effective_date': '2025-07-01', 'status': 'Active', 'docket': 'NJ BPU Docket ER23120924',
     'min_load_mw': 0.5, 'peak_demand_charge': 9.20, 'off_peak_demand_charge': 4.10,
     'energy_rate_peak': 0.055, 'energy_rate_off_peak': 0.040,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'PSEG Electric Tariff B.P.U.N.J. No. 17',
     'source_url': 'https://nj.pseg.com/aboutpseg/regulatorypage/electrictariffs',
     'page_reference': 'Schedule LPL'},

    {'utility': 'BGE', 'state': 'MD', 'region': 'Mid-Atlantic', 'iso_rto': 'PJM',
     'tariff_name': 'General Time-of-Use', 'rate_schedule': 'Schedule GT LV/TM-RT',
     'effective_date': '2025-06-01', 'status': 'Active', 'docket': 'MD PSC Case No. 9820',
     'min_load_mw': 0.025, 'peak_demand_charge': 8.50, 'off_peak_demand_charge': 3.80,
     'energy_rate_peak': 0.048, 'energy_rate_off_peak': 0.035,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'BGE Electric Rates and Service Tariff',
     'source_url': 'https://supplier.bge.com/electric/tariffs/index.asp',
     'page_reference': 'Schedule GT LV'},

    {'utility': 'ComEd (Exelon)', 'state': 'IL', 'region': 'Midwest', 'iso_rto': 'PJM',
     'tariff_name': 'Large Load Service', 'rate_schedule': 'Schedule 700 / TSA',
     'effective_date': '2026-01-06', 'status': 'Active', 'docket': 'ICC Dockets 25-0677/25-0678/25-0679',
     'min_load_mw': 50.0, 'peak_demand_charge': 10.50, 'off_peak_demand_charge': 4.80,
     'energy_rate_peak': 0.065, 'energy_rate_off_peak': 0.048,
     'contract_term_years': 10, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'ComEd TSA Requirements',
     'source_url': 'https://www.comed.com/current-rates-tariffs',
     'page_reference': 'Schedule 700'},

    # ==================== MIDWEST ====================
    {'utility': 'Ameren Missouri', 'state': 'MO', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Large Primary Service', 'rate_schedule': 'Schedule 11(M)',
     'effective_date': '2025-12-04', 'status': 'Active', 'docket': 'MO PSC Docket ER-2024-0319',
     'min_load_mw': 75.0, 'peak_demand_charge': 8.50, 'off_peak_demand_charge': 3.20,
     'energy_rate_peak': 0.038, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 12, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Ameren Missouri Large Load Customer Rate Plan',
     'source_url': 'https://s21.q4cdn.com/448935352/files/doc_presentations/2025/Nov/24/Ameren-Missouri-Large-Load-Customer-Rate-Plan-vfinal.pdf',
     'page_reference': 'Rate Plan, Nov 2025'},

    {'utility': 'Ameren Illinois', 'state': 'IL', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Large General Delivery', 'rate_schedule': 'Schedule DS-4',
     'effective_date': '2025-08-28', 'status': 'Active', 'docket': 'ICC Docket No. 25-0083',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.80, 'off_peak_demand_charge': 3.40,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Ameren Illinois DS-4 Tariff',
     'source_url': 'https://www.ameren.com/-/media/rates/illinois/non-residential/electric-rates/distribution-delivery/aiel14rtds4.ashx',
     'page_reference': 'Schedule DS-4'},

    {'utility': 'Evergy', 'state': 'KS/MO', 'region': 'Midwest', 'iso_rto': 'SPP',
     'tariff_name': 'Large Load Power Service', 'rate_schedule': 'Schedule LLPS',
     'effective_date': '2025-11-06', 'status': 'Active', 'docket': 'KS Docket 25-EKME-315-TAR',
     'min_load_mw': 75.0, 'peak_demand_charge': 7.20, 'off_peak_demand_charge': 2.80,
     'energy_rate_peak': 0.035, 'energy_rate_off_peak': 0.025,
     'contract_term_years': 17, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Evergy Large Load Power Service Tariff',
     'source_url': 'https://www.evergy.com/-/media/documents/billing/missouri/detailed_tariffs_mo/mo-west/large-power-service.pdf',
     'page_reference': 'Schedule LLPS'},

    {'utility': 'Consumers Energy', 'state': 'MI', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'General Primary Demand', 'rate_schedule': 'Schedule GPD',
     'effective_date': '2025-11-06', 'status': 'Active', 'docket': 'MI MPSC Docket U-21859',
     'min_load_mw': 100.0, 'peak_demand_charge': 9.80, 'off_peak_demand_charge': 4.20,
     'energy_rate_peak': 0.042, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 15, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'MPSC U-21859 Issue Brief',
     'source_url': 'https://www.michigan.gov/mpsc/-/media/Project/Websites/mpsc/consumer/info/briefs/Issue_Brief_U-21859_Consumers_Energy.pdf',
     'page_reference': 'Issue Brief, Nov 2025'},

    {'utility': 'DTE Energy', 'state': 'MI', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Primary Supply Agreement', 'rate_schedule': 'Schedule D11',
     'effective_date': '2025-12-18', 'status': 'Active', 'docket': 'MI MPSC Docket U-21990',
     'min_load_mw': 1.0, 'peak_demand_charge': 6.32, 'off_peak_demand_charge': 1.73,
     'energy_rate_peak': 0.038, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 19, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'DTE Primary Supply Agreement D11',
     'source_url': 'https://www.dteenergy.com/content/dam/dteenergy/deg/website/business/service-and-price/pricing/rate-options/PrimarySupplyAgreementD11.pdf',
     'page_reference': 'Schedule D11'},

    {'utility': 'We Energies', 'state': 'WI', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Large Load Service', 'rate_schedule': 'Proposed Schedule',
     'effective_date': 'TBD', 'status': 'Proposed', 'docket': 'WI PSC Docket 6630-FR-2024',
     'min_load_mw': 500.0, 'peak_demand_charge': 305.00, 'off_peak_demand_charge': 150.00,
     'energy_rate_peak': 0.055, 'energy_rate_off_peak': 0.040,
     'contract_term_years': 10, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'We Energies Data Center Rate Proposal',
     'source_url': 'https://www.we-energies.com/pdfs/etariffs/wisconsin/2025-rates-brochures.pdf',
     'page_reference': 'Proposed Large Load'},

    {'utility': 'Xcel Energy (MN)', 'state': 'MN', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Rate Book Schedule',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'MN PUC Dockets 24-320, 24-321',
     'min_load_mw': 1.0, 'peak_demand_charge': 8.90, 'off_peak_demand_charge': 3.60,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Xcel Energy Minnesota Rate Book',
     'source_url': 'https://www.xcelenergy.com/company/rates_and_regulations/rates/rate_books',
     'page_reference': 'MN Rate Book'},

    {'utility': 'MidAmerican Energy', 'state': 'IA', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Large Load Service', 'rate_schedule': 'Schedule 8',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'IUC Filing',
     'min_load_mw': 1.0, 'peak_demand_charge': 8.20, 'off_peak_demand_charge': 3.40,
     'energy_rate_peak': 0.042, 'energy_rate_off_peak': 0.030,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'MidAmerican Energy Tariff Schedule 8',
     'source_url': 'https://www.midamericanenergy.com/rates-tariffs',
     'page_reference': 'Schedule 8'},

    {'utility': 'Alliant Energy', 'state': 'WI/IA', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Schedule 760',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'PSCW/IUC Filing',
     'min_load_mw': 1.0, 'peak_demand_charge': 7.50, 'off_peak_demand_charge': 3.20,
     'energy_rate_peak': 0.040, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Alliant Energy Schedule 760',
     'source_url': 'https://www.alliantenergy.com/account-and-billing/understanding-bill-rates/iowa/electric',
     'page_reference': 'Schedule 760'},

    {'utility': 'Indianapolis Power & Light (AES)', 'state': 'IN', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'High Load Factor Service', 'rate_schedule': 'Rate HL',
     'effective_date': '2024-05-09', 'status': 'Active', 'docket': 'IURC Cause No. 45911',
     'min_load_mw': 2.0, 'peak_demand_charge': 24.09, 'off_peak_demand_charge': 10.50,
     'energy_rate_peak': 0.055, 'energy_rate_off_peak': 0.040,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'AES Indiana Rate HL Tariff',
     'source_url': 'https://www.aesindiana.com/sites/aesvault.com/files/2024-05/Rate-HL-High-Load-Factor--Primary-Distrib-Sub-Trans-and-Trans-Voltages-45911-Effective-05-09-24.pdf',
     'page_reference': 'Rate HL'},

    {'utility': 'Duke Energy Indiana', 'state': 'IN', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'High Load Factor Service', 'rate_schedule': 'Rate HLF',
     'effective_date': '2025-02-27', 'status': 'Active', 'docket': 'IURC Filing',
     'min_load_mw': 1.0, 'peak_demand_charge': 9.50, 'off_peak_demand_charge': 4.20,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Duke Energy Indiana Rate Tariff',
     'source_url': 'https://www.duke-energy.com/business/billing/rates',
     'page_reference': 'Rate HLF'},

    {'utility': 'NIPSCO', 'state': 'IN', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'General Service Large', 'rate_schedule': 'Rate 624/631',
     'effective_date': '2025-09-02', 'status': 'Active', 'docket': 'IURC Settlement June 2025',
     'min_load_mw': 0.05, 'peak_demand_charge': 35.83, 'off_peak_demand_charge': 15.20,
     'energy_rate_peak': 0.058, 'energy_rate_off_peak': 0.042,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'NIPSCO Rate 624 Tariff',
     'source_url': 'https://www.nipsco.com/docs/librariesprovider11/rates-and-tariffs/electric-rates/2025-to-current/rate-624.pdf',
     'page_reference': 'Rate 624'},

    {'utility': 'OPPD', 'state': 'NE', 'region': 'Midwest', 'iso_rto': 'SPP',
     'tariff_name': 'Industrial Service', 'rate_schedule': 'Industrial Tiers',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'Board Approval',
     'min_load_mw': 1.0, 'peak_demand_charge': 14.36, 'off_peak_demand_charge': 6.00,
     'energy_rate_peak': 0.0483, 'energy_rate_off_peak': 0.035,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'OPPD Industrial Rates',
     'source_url': 'https://www.oppd.com/business/business-rates/',
     'page_reference': 'Industrial Schedule'},

    {'utility': 'LG&E/KU (PPL)', 'state': 'KY', 'region': 'Midwest', 'iso_rto': 'MISO',
     'tariff_name': 'Extremely High Load Factor', 'rate_schedule': 'Schedule EHLF',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'KY PSC 2025-00114',
     'min_load_mw': 100.0, 'peak_demand_charge': 15.00, 'off_peak_demand_charge': 6.50,
     'energy_rate_peak': 0.048, 'energy_rate_off_peak': 0.035,
     'contract_term_years': 15, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Kentucky Utilities Company Tariff',
     'source_url': 'http://psc.ky.gov/tariffs/Electric/Kentucky%20Utilities%20Company/Tariff.pdf',
     'page_reference': 'Schedule EHLF'},

    {'utility': 'East Kentucky Power Coop', 'state': 'KY', 'region': 'Midwest', 'iso_rto': 'PJM',
     'tariff_name': 'Data Center Power', 'rate_schedule': 'Schedule DCP',
     'effective_date': '2025-10-31', 'status': 'Active', 'docket': 'KY PSC 2025-00140',
     'min_load_mw': 15.0, 'peak_demand_charge': 12.50, 'off_peak_demand_charge': 5.50,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 10, 'min_demand_pct': 60, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'EKPC Data Center Power Tariff',
     'source_url': 'https://psc.ky.gov/pscscf/2025%20Cases/2025-00140/20251030_PSC_ORDER.pdf',
     'page_reference': 'PSC Order 10/30/2025'},

    # ==================== MOUNTAIN WEST ====================
    {'utility': 'Black Hills Energy (SD)', 'state': 'SD', 'region': 'Mountain West', 'iso_rto': 'None',
     'tariff_name': 'Blockchain Interruptible', 'rate_schedule': 'Economic Flexible Load',
     'effective_date': '2026-01-28', 'status': 'Active', 'docket': 'SD PUC Docket EL25-019',
     'min_load_mw': 10.0, 'peak_demand_charge': 0, 'off_peak_demand_charge': 0,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.045,
     'contract_term_years': 2, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Black Hills Energy Blockchain Power Tariff',
     'source_url': 'https://puc.sd.gov/dockets/Electric/2025/default.aspx',
     'page_reference': 'Docket EL25-019'},

    {'utility': 'Black Hills Energy (WY)', 'state': 'WY', 'region': 'Mountain West', 'iso_rto': 'None',
     'tariff_name': 'Large Power Contract', 'rate_schedule': 'Schedule LPC',
     'effective_date': '2019-01-01', 'status': 'Active', 'docket': 'WY PSC Docket 20000-XXX',
     'min_load_mw': 13.0, 'peak_demand_charge': 0, 'off_peak_demand_charge': 0,
     'energy_rate_peak': 0.042, 'energy_rate_off_peak': 0.042,
     'contract_term_years': 3, 'min_demand_pct': 0, 'demand_ratchet': False,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Black Hills Wyoming Large Power Contract',
     'source_url': 'https://www.blackhillsenergy.com/billing-and-payments/rates-and-regulatory-information/wyoming-rates-and-regulatory-information',
     'page_reference': 'Schedule LPC'},

    {'utility': 'NV Energy', 'state': 'NV', 'region': 'West', 'iso_rto': 'None',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Schedule LGS-1',
     'effective_date': '2024-07-01', 'status': 'Active', 'docket': 'NV PUC Docket 24-XXXX',
     'min_load_mw': 1.0, 'peak_demand_charge': 8.50, 'off_peak_demand_charge': 4.20,
     'energy_rate_peak': 0.055, 'energy_rate_off_peak': 0.035,
     'contract_term_years': 10, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'NV Energy Schedule LGS-1',
     'source_url': 'https://www.nvenergy.com/publish/content/dam/nvenergy/brochures_702/handbook.pdf',
     'page_reference': 'Schedule LGS-1'},

    {'utility': 'Arizona Public Service (APS)', 'state': 'AZ', 'region': 'Southwest', 'iso_rto': 'None',
     'tariff_name': 'Large General Service TOU', 'rate_schedule': 'Schedule E-32',
     'effective_date': '2024-06-01', 'status': 'Active', 'docket': 'AZ CC Docket E-01345A-19-XXXX',
     'min_load_mw': 3.0, 'peak_demand_charge': 9.80, 'off_peak_demand_charge': 3.20,
     'energy_rate_peak': 0.048, 'energy_rate_off_peak': 0.028,
     'contract_term_years': 5, 'min_demand_pct': 75, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'APS E-32 Rate Schedule',
     'source_url': 'https://www.aps.com/en/Residential/Service-Plans/Compare-Service-Plans',
     'page_reference': 'Schedule E-32'},

    {'utility': 'Salt River Project (SRP)', 'state': 'AZ', 'region': 'Southwest', 'iso_rto': 'None',
     'tariff_name': 'Large Industrial Service', 'rate_schedule': 'Schedule E-65',
     'effective_date': '2025-11-01', 'status': 'Active', 'docket': 'SRP Board FY2026',
     'min_load_mw': 3.0, 'peak_demand_charge': 11.50, 'off_peak_demand_charge': 4.80,
     'energy_rate_peak': 0.058, 'energy_rate_off_peak': 0.038,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'SRP FY26 Ratebooks',
     'source_url': 'https://www.srpnet.com/assets/srpnet/pdf/price-plans/FY26/',
     'page_reference': 'Schedule E-65'},

    {'utility': 'Tucson Electric Power', 'state': 'AZ', 'region': 'Southwest', 'iso_rto': 'None',
     'tariff_name': 'Large Power Service', 'rate_schedule': 'Schedule 302 LPS',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'AZ CC Filing',
     'min_load_mw': 5.0, 'peak_demand_charge': 25.43, 'off_peak_demand_charge': 12.54,
     'energy_rate_peak': 0.065, 'energy_rate_off_peak': 0.045,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'TEP Tariff Schedule 302',
     'source_url': 'https://docs.tep.com/wp-content/uploads/302-TILPSTHV.pdf',
     'page_reference': 'Schedule 302'},

    {'utility': 'PNM (New Mexico)', 'state': 'NM', 'region': 'Southwest', 'iso_rto': 'None',
     'tariff_name': 'Large Power Service', 'rate_schedule': 'Schedule 35B',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'NM PRC Filing',
     'min_load_mw': 3.0, 'peak_demand_charge': 12.80, 'off_peak_demand_charge': 5.50,
     'energy_rate_peak': 0.055, 'energy_rate_off_peak': 0.038,
     'contract_term_years': 1, 'min_demand_pct': 50, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'PNM Schedule 35B',
     'source_url': 'https://www.pnm.com/documents/d/pnm.com/4th-revised-rate-35b',
     'page_reference': 'Schedule 35B'},

    {'utility': 'Xcel Energy (CO)', 'state': 'CO', 'region': 'Mountain West', 'iso_rto': 'None',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Schedule SG',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'CO PUC Proceeding',
     'min_load_mw': 1.0, 'peak_demand_charge': 10.20, 'off_peak_demand_charge': 4.50,
     'energy_rate_peak': 0.052, 'energy_rate_off_peak': 0.035,
     'contract_term_years': 15, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'Xcel Energy Colorado Schedule SG',
     'source_url': 'https://www.xcelenergy.com/company/rates_and_regulations/rates/rate_books',
     'page_reference': 'CO Rate Book'},

    {'utility': 'Colorado Springs Utilities', 'state': 'CO', 'region': 'Mountain West', 'iso_rto': 'None',
     'tariff_name': 'Large Load Service Agreement', 'rate_schedule': 'LLSA',
     'effective_date': '2026-01-01', 'status': 'Active', 'docket': 'CSU 2026 Rate Case',
     'min_load_mw': 10.0, 'peak_demand_charge': 15.50, 'off_peak_demand_charge': 6.80,
     'energy_rate_peak': 0.058, 'energy_rate_off_peak': 0.040,
     'contract_term_years': 10, 'min_demand_pct': 100, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'CSU 2026 Rate Case Filing',
     'source_url': 'https://www.csu.org/hubfs/Document-Library/2026-Rate-Case-Filing.pdf',
     'page_reference': 'LLSA Section'},

    {'utility': 'NorthWestern Energy', 'state': 'MT', 'region': 'Mountain West', 'iso_rto': 'None',
     'tariff_name': 'Large Load Tariff', 'rate_schedule': 'Proposed Schedule',
     'effective_date': 'TBD', 'status': 'Proposed', 'docket': 'MT PSC Sept 2025 Filing',
     'min_load_mw': 50.0, 'peak_demand_charge': 12.00, 'off_peak_demand_charge': 5.20,
     'energy_rate_peak': 0.052, 'energy_rate_off_peak': 0.038,
     'contract_term_years': 10, 'min_demand_pct': 80, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': True, 'exit_fee': True, 'credit_requirements': True, 'dc_specific': True,
     'source_document': 'NorthWestern Energy Large Load Filing',
     'source_url': 'https://northwesternenergy.com/billing-payment/rates-tariffs/rates-tariffs-montana/montana-rate-review/',
     'page_reference': 'Large Load Tariff'},

    {'utility': 'Rocky Mountain Power (PacifiCorp)', 'state': 'UT/ID/WY', 'region': 'Mountain West', 'iso_rto': 'None',
     'tariff_name': 'Large General Service', 'rate_schedule': 'Schedule 31',
     'effective_date': '2025-11-01', 'status': 'Active', 'docket': 'UT PSC Docket 24-035-XX',
     'min_load_mw': 1.0, 'peak_demand_charge': 9.56, 'off_peak_demand_charge': 6.68,
     'energy_rate_peak': 0.048, 'energy_rate_off_peak': 0.035,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Rocky Mountain Power Schedule 31',
     'source_url': 'https://www.rockymountainpower.net/content/dam/pcorp/documents/en/rockymountainpower/rates-regulation/utah/rates/031_Partial_Requirements_Service_Large_General_Service_1000kW_and_Over.pdf',
     'page_reference': 'Schedule 31'},

    # ==================== WEST / CALIFORNIA ====================
    {'utility': 'Pacific Gas & Electric (PG&E)', 'state': 'CA', 'region': 'West', 'iso_rto': 'CAISO',
     'tariff_name': 'Large Power', 'rate_schedule': 'Schedule E-20',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'CPUC A.23-11-XXX',
     'min_load_mw': 1.0, 'peak_demand_charge': 22.50, 'off_peak_demand_charge': 8.40,
     'energy_rate_peak': 0.185, 'energy_rate_off_peak': 0.120,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'PG&E Schedule E-20 Large Power Service',
     'source_url': 'https://www.pge.com/tariffs/assets/pdf/tariffbook/ELEC_SCHEDS_E-20.pdf',
     'page_reference': 'Schedule E-20'},

    {'utility': 'Southern California Edison (SCE)', 'state': 'CA', 'region': 'West', 'iso_rto': 'CAISO',
     'tariff_name': 'Large TOU', 'rate_schedule': 'Schedule TOU-8',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'CPUC A.22-05-XXX',
     'min_load_mw': 0.5, 'peak_demand_charge': 18.80, 'off_peak_demand_charge': 6.20,
     'energy_rate_peak': 0.165, 'energy_rate_off_peak': 0.105,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'SCE TOU-8 Rate Fact Sheet',
     'source_url': 'https://www.sce.com/sites/default/files/inline-files/TOU-8%20Rate%20Fact%20Sheet_WCAG%20(1).pdf',
     'page_reference': 'TOU-8 Fact Sheet'},

    {'utility': 'San Diego Gas & Electric (SDG&E)', 'state': 'CA', 'region': 'West', 'iso_rto': 'CAISO',
     'tariff_name': 'Large C&I TOU', 'rate_schedule': 'Schedule AL-TOU',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'CPUC A.23-01-XXX',
     'min_load_mw': 0.02, 'peak_demand_charge': 15.20, 'off_peak_demand_charge': 5.80,
     'energy_rate_peak': 0.145, 'energy_rate_off_peak': 0.095,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'SDG&E Schedule AL-TOU',
     'source_url': 'https://www.sdge.com/sites/default/files/elec_elec-scheds_al-tou.pdf',
     'page_reference': 'Schedule AL-TOU'},

    {'utility': 'Portland General Electric', 'state': 'OR', 'region': 'West', 'iso_rto': 'None',
     'tariff_name': 'Large Industrial', 'rate_schedule': 'Schedule 89',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'OR PUC Docket UE 435',
     'min_load_mw': 4.0, 'peak_demand_charge': 8.20, 'off_peak_demand_charge': 3.40,
     'energy_rate_peak': 0.068, 'energy_rate_off_peak': 0.048,
     'contract_term_years': 10, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'PGE Schedule 89 Large Industrial',
     'source_url': 'https://portlandgeneral.com/about/info/rates-and-regulatory/tariff',
     'page_reference': 'Schedule 89'},

    {'utility': 'Idaho Power', 'state': 'ID', 'region': 'West', 'iso_rto': 'None',
     'tariff_name': 'Large Power Service', 'rate_schedule': 'Schedule 19',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'ID PUC Case',
     'min_load_mw': 1.0, 'peak_demand_charge': 10.50, 'off_peak_demand_charge': 8.45,
     'energy_rate_peak': 0.045, 'energy_rate_off_peak': 0.032,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Idaho Power Schedule 19',
     'source_url': 'https://docs.idahopower.com/pdfs/aboutus/ratesregulatory/tariffs/191.pdf',
     'page_reference': 'Schedule 19-1'},

    {'utility': 'Puget Sound Energy', 'state': 'WA', 'region': 'West', 'iso_rto': 'None',
     'tariff_name': 'Large Demand', 'rate_schedule': 'Schedule 26',
     'effective_date': '2025-01-29', 'status': 'Active', 'docket': 'WA UTC Docket',
     'min_load_mw': 0.35, 'peak_demand_charge': 7.80, 'off_peak_demand_charge': 3.20,
     'energy_rate_peak': 0.072, 'energy_rate_off_peak': 0.052,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'PSE Schedule 26',
     'source_url': 'https://www.pse.com/en/pages/rates/schedule-summaries',
     'page_reference': 'Schedule 26'},

    # ==================== NORTHEAST ====================
    {'utility': 'ConEdison', 'state': 'NY', 'region': 'Northeast', 'iso_rto': 'NYISO',
     'tariff_name': 'PASNY Delivery Service', 'rate_schedule': 'P.S.C. No. 12',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'NY PSC Case',
     'min_load_mw': 1.0, 'peak_demand_charge': 28.50, 'off_peak_demand_charge': 12.40,
     'energy_rate_peak': 0.165, 'energy_rate_off_peak': 0.110,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'ConEdison Electric Tariff P.S.C. No. 12',
     'source_url': 'https://www.coned.com/en/rates-tariffs/rates',
     'page_reference': 'P.S.C. No. 12'},

    {'utility': 'National Grid (NY)', 'state': 'NY', 'region': 'Northeast', 'iso_rto': 'NYISO',
     'tariff_name': 'Large Commercial Service', 'rate_schedule': 'Schedule LC',
     'effective_date': '2025-01-01', 'status': 'Active', 'docket': 'NY PSC Case',
     'min_load_mw': 0.1, 'peak_demand_charge': 12.80, 'off_peak_demand_charge': 5.20,
     'energy_rate_peak': 0.095, 'energy_rate_off_peak': 0.065,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'National Grid NY Schedule LC',
     'source_url': 'https://www.nationalgridus.com/Upstate-NY-Business/Rates/Service-Rates',
     'page_reference': 'Schedule LC'},

    {'utility': 'Eversource (CT)', 'state': 'CT', 'region': 'Northeast', 'iso_rto': 'ISO-NE',
     'tariff_name': 'Intermediate TOU General', 'rate_schedule': 'Schedule 37',
     'effective_date': '2025-07-01', 'status': 'Active', 'docket': 'CT PURA Docket',
     'min_load_mw': 0.35, 'peak_demand_charge': 14.20, 'off_peak_demand_charge': 6.80,
     'energy_rate_peak': 0.125, 'energy_rate_off_peak': 0.085,
     'contract_term_years': 5, 'min_demand_pct': 0, 'demand_ratchet': True,
     'ciac_required': True, 'take_or_pay': False, 'exit_fee': False, 'credit_requirements': True, 'dc_specific': False,
     'source_document': 'Eversource CT Schedule 37',
     'source_url': 'https://www.eversource.com/residential/account-billing/manage-bill/about-your-bill/rates-tariffs',
     'page_reference': 'Schedule 37'},
]

def create_workbook():
    """Create comprehensive Excel workbook."""
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    high_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    mid_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    low_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # Sheet 1: Complete Database
    ws1 = wb.active
    ws1.title = 'Tariff Database'

    headers = ['Utility', 'State', 'Region', 'ISO/RTO', 'Tariff Name', 'Rate Schedule',
               'Effective Date', 'Status', 'Docket', 'Min Load (MW)',
               'Peak Demand ($/kW)', 'Off-Peak Demand', 'Energy Peak ($/kWh)', 'Energy Off-Peak',
               'Contract (Years)', 'Min Demand %', 'Ratchet', 'CIAC', 'Take-or-Pay', 'Exit Fee',
               'DC Specific', 'Protection Score', 'Score Points', 'Blended Rate ($/kWh)',
               'Annual Cost ($M)', 'Source Document', 'Page Reference']

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row, t in enumerate(UTILITIES, 2):
        points, score = calculate_protection_score(t)
        blended = calculate_blended_rate(t)
        annual_cost = blended * MONTHLY_KWH * 12 / 1_000_000  # Convert to $M

        data = [
            t.get('utility'), t.get('state'), t.get('region'), t.get('iso_rto'),
            t.get('tariff_name'), t.get('rate_schedule'), t.get('effective_date'), t.get('status'),
            t.get('docket'), t.get('min_load_mw'),
            t.get('peak_demand_charge'), t.get('off_peak_demand_charge'),
            t.get('energy_rate_peak'), t.get('energy_rate_off_peak'),
            t.get('contract_term_years'), t.get('min_demand_pct'),
            'Yes' if t.get('demand_ratchet') else 'No',
            'Yes' if t.get('ciac_required') else 'No',
            'Yes' if t.get('take_or_pay') else 'No',
            'Yes' if t.get('exit_fee') else 'No',
            'Yes' if t.get('dc_specific') else 'No',
            score, points, round(blended, 5), round(annual_cost, 1),
            t.get('source_document'), t.get('page_reference')
        ]

        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = border
            if col == 22:  # Protection Score
                if val == 'High': cell.fill = high_fill
                elif val == 'Mid': cell.fill = mid_fill
                else: cell.fill = low_fill
            if col in [24, 25]:  # Blended rate and annual cost
                cell.number_format = '0.00000' if col == 24 else '#,##0.0'

    # Set widths
    widths = [30, 8, 12, 8, 22, 20, 12, 10, 22, 8, 10, 10, 10, 10, 8, 8, 6, 6, 6, 6, 6, 10, 8, 12, 12, 35, 25]
    for col, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A2'

    # Sheet 2: Blended Rate Summary
    ws2 = wb.create_sheet('Blended Rate Analysis')

    headers2 = ['Utility', 'State', 'Region', 'Blended Rate ($/kWh)', 'Annual Cost ($M)',
                'Monthly Demand Cost', 'Monthly Energy Cost', 'Protection Score', 'Contract Years']

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Sort by blended rate
    sorted_utils = sorted(UTILITIES, key=lambda x: calculate_blended_rate(x))

    for row, t in enumerate(sorted_utils, 2):
        blended = calculate_blended_rate(t)
        annual = blended * MONTHLY_KWH * 12 / 1_000_000

        # Calculate components
        peak_demand = t.get('peak_demand_charge', 0) or 0
        offpeak_demand = t.get('off_peak_demand_charge', 0) or 0
        energy_peak = t.get('energy_rate_peak', 0) or 0
        energy_offpeak = t.get('energy_rate_off_peak', 0) or energy_peak * 0.7

        demand_cost = (peak_demand * BILLING_DEMAND_KW) + (offpeak_demand * BILLING_DEMAND_KW * 0.5)
        energy_cost = (energy_peak * PEAK_KWH) + (energy_offpeak * OFFPEAK_KWH)

        _, score = calculate_protection_score(t)

        data = [t.get('utility'), t.get('state'), t.get('region'),
                round(blended, 5), round(annual, 1),
                round(demand_cost / 1_000_000, 2), round(energy_cost / 1_000_000, 2),
                score, t.get('contract_term_years')]

        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = border

    widths2 = [30, 8, 12, 15, 12, 15, 15, 12, 10]
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sheet 3: Protection Matrix
    ws3 = wb.create_sheet('Protection Matrix')

    headers3 = ['Utility', 'State', 'Min Demand %', 'Contract Yrs', 'CIAC', 'Take-or-Pay',
                'Exit Fee', 'Ratchet', 'Credit Req', 'DC Specific', 'Score', 'Rating']

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Sort by protection score (descending)
    sorted_by_prot = sorted(UTILITIES, key=lambda x: calculate_protection_score(x)[0], reverse=True)

    for row, t in enumerate(sorted_by_prot, 2):
        points, score = calculate_protection_score(t)

        data = [t.get('utility'), t.get('state'),
                t.get('min_demand_pct', 0), t.get('contract_term_years', 0),
                'Yes' if t.get('ciac_required') else 'No',
                'Yes' if t.get('take_or_pay') else 'No',
                'Yes' if t.get('exit_fee') else 'No',
                'Yes' if t.get('demand_ratchet') else 'No',
                'Yes' if t.get('credit_requirements') else 'No',
                'Yes' if t.get('dc_specific') else 'No',
                points, score]

        for col, val in enumerate(data, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.border = border
            if col == 12:
                if val == 'High': cell.fill = high_fill
                elif val == 'Mid': cell.fill = mid_fill
                else: cell.fill = low_fill

    widths3 = [30, 8, 10, 10, 8, 10, 8, 8, 10, 10, 8, 8]
    for col, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # Save
    path = '/sessions/laughing-peaceful-archimedes/mnt/power-insight/Large_Load_Tariff_Database_Comprehensive.xlsx'
    wb.save(path)

    print(f"Database created: {path}")
    print(f"Total utilities: {len(UTILITIES)}")

    # Stats
    high = sum(1 for t in UTILITIES if calculate_protection_score(t)[1] == 'High')
    mid = sum(1 for t in UTILITIES if calculate_protection_score(t)[1] == 'Mid')
    low = sum(1 for t in UTILITIES if calculate_protection_score(t)[1] == 'Low')
    print(f"High Protection: {high}, Mid: {mid}, Low: {low}")

    rates = [calculate_blended_rate(t) for t in UTILITIES]
    print(f"Blended Rate Range: ${min(rates):.5f} - ${max(rates):.5f}/kWh")
    print(f"Average Blended Rate: ${sum(rates)/len(rates):.5f}/kWh")

if __name__ == '__main__':
    create_workbook()
